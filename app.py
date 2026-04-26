from flask import Flask, render_template, redirect, url_for, request, flash, Response, send_file
from flask_login import LoginManager, login_user, login_required, logout_user, UserMixin, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import sqlite3
from datetime import datetime, timedelta
from email.message import EmailMessage
import csv
import io
import os
from pathlib import Path
import smtplib

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads" / "pto_documents"
OVERTIME_UPLOAD_DIR = BASE_DIR / "uploads" / "overtime_schedules"
DEFAULT_SECRET_KEY = "change-me-before-production"

app = Flask(__name__)
app.secret_key = os.getenv("PTO_SECRET_KEY", DEFAULT_SECRET_KEY)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

DB = os.getenv("PTO_DB_PATH", str(BASE_DIR / "database.db"))
MAX_OFFICERS_OFF = 2
ROTATION_START_DATE = "2026-04-19"
ANCHOR_ACTIVE_TEAMS = {"Black", "Blue"}
TEAM_OPTIONS = ("Red", "Black", "Gold", "Blue", "Command Staff")
SITE_OPTIONS = ("Site Alpha", "Site Beta")
BUILDING_OPTIONS = (
    "Operations Center",
    "Primary Gatehouse",
    "North Campus Lobby",
    "South Campus Lobby",
    "Logistics Hub",
    "Research Annex",
    "Security Control Center",
    "Client Site A",
    "Client Site B",
    "Client Site C",
    "Command Staff",
)
RANK_OPTIONS = ("Officer", "Sergeant", "Lieutenant", "Training Lieutenant", "Captain", "Director")
SUPERVISOR_RANKS = ("Lieutenant", "Training Lieutenant", "Captain", "Director")
COMMAND_REVIEW_RANKS = ("Captain", "Director")
LATE_PUNCH_THRESHOLD_MINUTES = 15
UNPAID_BREAK_MINUTES = 20
PAID_LUNCH_MINUTES = 45
COMMAND_TEAM_RANKS = ("Training Lieutenant", "Captain", "Director")
SPECIAL_PTO_TYPES = ("Bereavement", "Jury Duty")
ALLOWED_DOCUMENT_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
ALLOWED_SCHEDULE_EXTENSIONS = {".xlsx"}
MAX_OT_HOURS_PER_WEEK = 20.0
MAX_OT_HOURS_PER_PAY_PERIOD = 40.0

# ---------------- DATABASE ----------------
def init_db():
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        password TEXT,
        role TEXT,
        rank TEXT DEFAULT 'Officer',
        email TEXT,
        phone TEXT,
        team TEXT DEFAULT '',
        site TEXT DEFAULT '',
        building TEXT DEFAULT '',
        first_name TEXT DEFAULT '',
        last_name TEXT DEFAULT ''
    )''')

    c.execute("PRAGMA table_info(users)")
    user_columns = {row[1] for row in c.fetchall()}
    if "team" not in user_columns:
        c.execute("ALTER TABLE users ADD COLUMN team TEXT DEFAULT ''")
    if "site" not in user_columns:
        c.execute("ALTER TABLE users ADD COLUMN site TEXT DEFAULT ''")
    if "building" not in user_columns:
        c.execute("ALTER TABLE users ADD COLUMN building TEXT DEFAULT ''")
    if "first_name" not in user_columns:
        c.execute("ALTER TABLE users ADD COLUMN first_name TEXT DEFAULT ''")
    if "last_name" not in user_columns:
        c.execute("ALTER TABLE users ADD COLUMN last_name TEXT DEFAULT ''")

    c.execute('''CREATE TABLE IF NOT EXISTS overtime_schedule_uploads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_label TEXT,
        site TEXT,
        shift TEXT,
        original_filename TEXT,
        file_path TEXT,
        uploaded_by INTEGER,
        uploaded_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS overtime_openings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_label TEXT,
        site TEXT,
        shift TEXT,
        location TEXT,
        opening_date TEXT,
        day_name TEXT,
        slot_label TEXT,
        priority TEXT,
        shift_start TEXT,
        shift_end TEXT,
        source_upload_id INTEGER,
        status TEXT DEFAULT 'Open',
        assigned_user_id INTEGER,
        assigned_at TEXT,
        assigned_by TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS overtime_baseline_hours (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_label TEXT,
        site TEXT,
        shift TEXT,
        roster_name TEXT,
        week_1_hours REAL DEFAULT 0,
        week_2_hours REAL DEFAULT 0,
        pay_period_hours REAL DEFAULT 0,
        uploaded_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS overtime_baseline_shifts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_label TEXT,
        site TEXT,
        shift TEXT,
        roster_name TEXT,
        work_date TEXT,
        location TEXT,
        uploaded_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS overtime_applications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        opening_id INTEGER,
        user_id INTEGER,
        note TEXT DEFAULT '',
        status TEXT DEFAULT 'Applied',
        submitted_at TEXT
    )''')


    c.execute('''CREATE TABLE IF NOT EXISTS pto_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        start_date TEXT,
        end_date TEXT,
        status TEXT DEFAULT 'Pending',
        admin_note TEXT,
        created_at TEXT,
        request_kind TEXT DEFAULT 'PTO',
        special_type TEXT DEFAULT '',
        documentation_path TEXT DEFAULT '',
        override_approved INTEGER DEFAULT 0
    )''')

    c.execute("PRAGMA table_info(pto_requests)")
    pto_columns = {row[1] for row in c.fetchall()}
    if "admin_note" not in pto_columns:
        c.execute("ALTER TABLE pto_requests ADD COLUMN admin_note TEXT DEFAULT ''")
    if "request_kind" not in pto_columns:
        c.execute("ALTER TABLE pto_requests ADD COLUMN request_kind TEXT DEFAULT 'PTO'")
    if "special_type" not in pto_columns:
        c.execute("ALTER TABLE pto_requests ADD COLUMN special_type TEXT DEFAULT ''")
    if "documentation_path" not in pto_columns:
        c.execute("ALTER TABLE pto_requests ADD COLUMN documentation_path TEXT DEFAULT ''")
    if "override_approved" not in pto_columns:
        c.execute("ALTER TABLE pto_requests ADD COLUMN override_approved INTEGER DEFAULT 0")

    c.execute('''CREATE TABLE IF NOT EXISTS shift_swaps (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        requesting_user_id INTEGER,
        swap_with_user_id INTEGER,
        start_date TEXT,
        end_date TEXT,
        reason TEXT,
        status TEXT DEFAULT 'Pending',
        approved_by TEXT,
        review_note TEXT DEFAULT '',
        created_at TEXT
    )''')

    c.execute("PRAGMA table_info(shift_swaps)")
    shift_swap_columns = {row[1] for row in c.fetchall()}
    if "review_note" not in shift_swap_columns:
        c.execute("ALTER TABLE shift_swaps ADD COLUMN review_note TEXT DEFAULT ''")

    c.execute('''CREATE TABLE IF NOT EXISTS time_clock_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        clock_date TEXT,
        clock_in TEXT,
        clock_out TEXT,
        worked_hours REAL,
        expected_hours REAL,
        created_at TEXT,
        updated_at TEXT,
        team TEXT,
        schedule_note TEXT
    )''')

    c.execute("PRAGMA table_info(time_clock_entries)")
    time_clock_columns = {row[1] for row in c.fetchall()}
    if "team" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN team TEXT")
    if "schedule_note" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN schedule_note TEXT")
    if "supervisor_note" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN supervisor_note TEXT DEFAULT ''")
    if "corrected_by" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN corrected_by TEXT DEFAULT ''")
    if "corrected_at" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN corrected_at TEXT")
    if "entry_source" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN entry_source TEXT DEFAULT 'clock'")
    if "break1_start" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN break1_start TEXT")
    if "break1_end" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN break1_end TEXT")
    if "break2_start" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN break2_start TEXT")
    if "break2_end" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN break2_end TEXT")
    if "lunch_start" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN lunch_start TEXT")
    if "lunch_end" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN lunch_end TEXT")
    if "unpaid_break_minutes" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN unpaid_break_minutes REAL DEFAULT 0")
    if "paid_lunch_minutes" not in time_clock_columns:
        c.execute("ALTER TABLE time_clock_entries ADD COLUMN paid_lunch_minutes REAL DEFAULT 0")

    c.execute('''CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        title TEXT,
        message TEXT,
        category TEXT DEFAULT 'info',
        link TEXT DEFAULT '',
        is_read INTEGER DEFAULT 0,
        created_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entity_type TEXT,
        entity_id INTEGER,
        action TEXT,
        actor_user_id INTEGER,
        actor_username TEXT,
        target_user_id INTEGER,
        details TEXT,
        created_at TEXT
    )''')

    conn.commit()
    conn.close()

init_db()

def get_team_metadata(team):
    metadata = {
        "Red": {"label": "Red", "color": "#9b1c31", "background": "#fdecef", "shift_label": "Days"},
        "Black": {"label": "Black", "color": "#1f2933", "background": "#e5e7eb", "shift_label": "Days"},
        "Gold": {"label": "Gold", "color": "#9a6700", "background": "#fff3c4", "shift_label": "Nights"},
        "Blue": {"label": "Blue", "color": "#0f4c81", "background": "#dbeafe", "shift_label": "Nights"},
        "Command Staff": {"label": "Command Staff", "color": "#5b21b6", "background": "#ede9fe", "shift_label": "Command"},
    }
    default = {"label": team or "Unassigned", "color": "#4b5563", "background": "#e5e7eb", "shift_label": "Unassigned"}
    return metadata.get(team, default)


def get_team_schedule(team, clock_date_str):
    if team == "Command Staff":
        return {
            "team": team,
            "rotation_week": None,
            "expected_hours": None,
            "schedule_note": "Command staff schedule",
            "is_scheduled": False,
            "team_meta": get_team_metadata(team),
        }

    if not team or team not in TEAM_OPTIONS:
        return {
            "team": team or "",
            "rotation_week": None,
            "expected_hours": None,
            "schedule_note": "Team not assigned",
            "is_scheduled": False,
            "team_meta": get_team_metadata(team),
        }

    anchor_date = datetime.strptime(ROTATION_START_DATE, "%Y-%m-%d").date()
    clock_date = datetime.strptime(clock_date_str, "%Y-%m-%d").date()
    rotation_week = ((clock_date - anchor_date).days // 7) % 2 + 1

    active_teams = ANCHOR_ACTIVE_TEAMS if rotation_week == 1 else {"Red", "Gold"}
    is_pattern_a = team in active_teams
    weekday = clock_date.weekday()

    if is_pattern_a:
        expected_hours = {0: 11.4, 1: 5.8, 2: 0.0, 3: 0.0, 4: 11.4, 5: 11.4, 6: 11.4}[weekday]
        schedule_note = {
            0: "Scheduled",
            1: "Tuesday first half",
            2: "Off",
            3: "Off",
            4: "Scheduled",
            5: "Scheduled",
            6: "Scheduled",
        }[weekday]
    else:
        expected_hours = {0: 0.0, 1: 5.8, 2: 11.4, 3: 11.4, 4: 0.0, 5: 0.0, 6: 0.0}[weekday]
        schedule_note = {
            0: "Off",
            1: "Tuesday second half",
            2: "Scheduled",
            3: "Scheduled",
            4: "Off",
            5: "Off",
            6: "Off",
        }[weekday]

    return {
        "team": team,
        "rotation_week": rotation_week,
        "expected_hours": expected_hours,
        "schedule_note": schedule_note,
        "is_scheduled": expected_hours > 0,
        "team_meta": get_team_metadata(team),
    }


def get_rotation_week_for_date(clock_date_str):
    anchor_date = datetime.strptime(ROTATION_START_DATE, "%Y-%m-%d").date()
    clock_date = datetime.strptime(clock_date_str, "%Y-%m-%d").date()
    return ((clock_date - anchor_date).days // 7) % 2 + 1


def get_rotation_shift_teams(clock_date_str, shift):
    rotation_week = get_rotation_week_for_date(clock_date_str)
    active_teams = ANCHOR_ACTIVE_TEAMS if rotation_week == 1 else {"Red", "Gold"}
    if shift == "day":
        active_team = "Black" if "Black" in active_teams else "Red"
        inactive_team = "Red" if active_team == "Black" else "Black"
    else:
        active_team = "Blue" if "Blue" in active_teams else "Gold"
        inactive_team = "Gold" if active_team == "Blue" else "Blue"
    return active_team, inactive_team


def get_opposite_ot_teams_for_user(team):
    if team in {"Black", "Blue"}:
        return ("Red", "Gold")
    if team in {"Red", "Gold"}:
        return ("Black", "Blue")
    return ()


def get_scheduled_team_for_opening(opening_date_str, shift, slot_label=""):
    if shift not in {"day", "night"}:
        return ""
    weekday = datetime.strptime(opening_date_str, "%Y-%m-%d").weekday()
    active_team, inactive_team = get_rotation_shift_teams(opening_date_str, shift)

    slot_label_text = str(slot_label or "").lower()
    is_split = "split" in slot_label_text

    if weekday == 1:
        return active_team if is_split else inactive_team

    if weekday in {6, 0, 4, 5}:
        return active_team
    return inactive_team


def enrich_overtime_opening_row(row):
    if isinstance(row, dict):
        opening = dict(row)
    elif hasattr(row, "keys"):
        opening = {key: row[key] for key in row.keys()}
    else:
        opening = {
            "id": row[0],
            "period_label": row[1],
            "site": row[2],
            "shift": row[3],
            "location": row[4],
            "opening_date": row[5],
            "day_name": row[6],
            "slot_label": row[7],
            "priority": row[8],
            "shift_start": row[9],
            "shift_end": row[10],
            "source_upload_id": row[11],
            "status": row[12],
            "assigned_user_id": row[13],
            "assigned_at": row[14],
            "assigned_by": row[15],
        }
    team = get_scheduled_team_for_opening(
        opening["opening_date"],
        opening["shift"],
        opening.get("slot_label", ""),
    )
    team_meta = get_team_metadata(team)
    opening["team"] = team
    opening["team_meta"] = team_meta
    opening["team_label"] = f"{team_meta['label']} {team_meta['shift_label']}" if team else opening["shift"].title()
    return opening


def filter_openings_for_user(openings, user):
    if user_is_command_team(user):
        return openings
    allowed_teams = get_opposite_ot_teams_for_user(getattr(user, "team", "") or "")
    if not allowed_teams:
        return openings
    return [opening for opening in openings if opening.get("team") in allowed_teams]


def calculate_duration_minutes(start_str, end_str, fallback_end_str=None):
    if not start_str:
        return 0
    end_value = end_str or fallback_end_str
    if not end_value:
        return 0
    start = datetime.fromisoformat(start_str)
    end = datetime.fromisoformat(end_value)
    seconds = max((end - start).total_seconds(), 0)
    return round(seconds / 60, 2)


def calculate_gross_hours(clock_in_str, clock_out_str):
    if not clock_in_str or not clock_out_str:
        return None
    clock_in = datetime.fromisoformat(clock_in_str)
    clock_out = datetime.fromisoformat(clock_out_str)
    seconds = max((clock_out - clock_in).total_seconds(), 0)
    return round(seconds / 3600, 2)


def calculate_worked_hours(clock_in_str, clock_out_str, unpaid_break_minutes=0):
    gross_hours = calculate_gross_hours(clock_in_str, clock_out_str)
    if gross_hours is None:
        return None
    net_minutes = max((gross_hours * 60) - float(unpaid_break_minutes or 0), 0)
    return round(net_minutes / 60, 2)


def get_allowed_unpaid_breaks(expected_hours):
    if expected_hours is None or expected_hours <= 0:
        return 0
    return 1 if expected_hours <= 6 else 2


def user_is_command_team(user):
    role = getattr(user, "role", "") or ""
    rank = getattr(user, "rank", "") or ""
    team = getattr(user, "team", "") or ""
    return role == "admin" or rank in COMMAND_TEAM_RANKS or team == "Command Staff"


def rank_team_is_command(rank, team, role=""):
    return (role or "") == "admin" or (rank or "") in COMMAND_TEAM_RANKS or (team or "") == "Command Staff"


def is_allowed_document(filename):
    suffix = Path(filename or "").suffix.lower()
    return suffix in ALLOWED_DOCUMENT_EXTENSIONS


def normalize_phone_number(value):
    digits = "".join(ch for ch in (value or "") if ch.isdigit())
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return value.strip()


def get_display_name_from_values(first_name, last_name, username):
    full_name = " ".join(part.strip() for part in [first_name or "", last_name or ""] if part and part.strip())
    return full_name or (username or "")


def get_roster_name_from_values(first_name, last_name, username):
    first = (first_name or "").strip()
    last = (last_name or "").strip()
    if first and last:
        first_token = first.split()[0]
        roster_prefix = first_token if len(first_token) <= 2 else first_token[0]
        return f"{roster_prefix}.{last}".replace(" ", "")
    return (username or "").strip()


def is_allowed_schedule_file(filename):
    suffix = Path(filename or "").suffix.lower()
    return suffix in ALLOWED_SCHEDULE_EXTENSIONS


def save_overtime_schedule(file_storage, period_label, site, shift):
    filename = secure_filename(file_storage.filename or "")
    if not filename:
        raise ValueError("Schedule file is missing a filename.")
    if not is_allowed_schedule_file(filename):
        raise ValueError("Schedule must be an XLSX file.")

    OVERTIME_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    safe_period = secure_filename(period_label or "current-period")
    safe_site = secure_filename(site or "site")
    saved_name = f"{safe_period}_{safe_site}_{shift}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
    destination = OVERTIME_UPLOAD_DIR / saved_name
    file_storage.save(destination)
    return str(destination)


def clean_schedule_name(value):
    return " ".join(str(value).replace("\n", " ").split())


def is_schedule_placeholder(value):
    text = clean_schedule_name(value).upper()
    if text in {"CRITICAL", "FLEX", "CLOSED", "OFF", "PTO", "OPEN AVAIL.", "OPEN AVAIL"}:
        return True
    return "FLEX" in text or "CRITICAL" in text or "OPEN AVAIL" in text


def is_schedule_opening(value):
    text = clean_schedule_name(value).upper()
    return "CRITICAL" in text or "FLEX" in text or "OPEN AVAIL" in text


def parse_schedule_time_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def calculate_schedule_hours(start_value, end_value):
    start = parse_schedule_time_value(start_value)
    end = parse_schedule_time_value(end_value)
    if not start or not end:
        return 0.0
    if end <= start:
        end += timedelta(days=1)
    raw_hours = round((end - start).total_seconds() / 3600, 2)
    if raw_hours <= 6.5:
        return 5.8
    return 11.4


def build_schedule_columns(ws, start_col):
    columns = []
    prior_date = None
    repeat_count = 0

    for col in range(start_col, start_col + 8):
        raw_date = ws.cell(2, col).value
        raw_day = ws.cell(3, col).value

        if raw_date:
            prior_date = raw_date
            repeat_count = 1
            date_text = raw_date.date().isoformat() if isinstance(raw_date, datetime) else str(raw_date)
            day_text = str(raw_day).strip() if raw_day else ""
            label = day_text
        else:
            if prior_date is None:
                continue
            repeat_count += 1
            date_text = prior_date.date().isoformat() if isinstance(prior_date, datetime) else str(prior_date)
            prior_day = columns[-1]["day_name"]
            day_text = str(raw_day).strip() if raw_day else f"{prior_day} Split"
            label = f"{prior_day} split {repeat_count}"

        columns.append(
            {
                "column": col,
                "date": date_text,
                "day_name": day_text or columns[-1]["day_name"],
                "label": label,
            }
        )
    return columns


def get_slot_schedule_values(shift, opening_date_str, slot_label, start_value, end_value):
    slot_label_text = (slot_label or "").lower()
    weekday = datetime.strptime(opening_date_str, "%Y-%m-%d").weekday()
    is_split = "split" in slot_label_text

    if weekday == 1:
        if shift == "day":
            if is_split:
                return "12:00:00", "18:00:00", 5.8
            return "06:00:00", "12:00:00", 5.8
        if shift == "night":
            if is_split:
                return "00:00:00", "06:00:00", 5.8
            return "18:00:00", "00:00:00", 5.8

    start_time = str(start_value) if start_value is not None else ""
    end_time = str(end_value) if end_value is not None else ""
    return start_time, end_time, calculate_schedule_hours(start_value, end_value)


def find_schedule_sections(ws, start_col):
    sections = []
    for row in range(1, ws.max_row + 1):
        marker = ws.cell(row, start_col - 2).value
        location = ws.cell(row, start_col).value
        if str(marker).strip() == "In" and location:
            sections.append((row, clean_schedule_name(location)))
    return sections


def parse_overtime_schedule_workbook(path, site, shift, period_label):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    week_specs = [(3, "week_1"), (13, "week_2")]

    openings = []
    baseline_hours = {}
    baseline_shifts = []

    for start_col, week_name in week_specs:
        columns = build_schedule_columns(ws, start_col)
        if not columns:
            continue
        sections = find_schedule_sections(ws, start_col)
        for idx, (_, location) in enumerate(sections):
            current_start_time = None
            current_end_time = None
            start_row = sections[idx][0] + 1
            end_row = sections[idx + 1][0] - 1 if idx + 1 < len(sections) else ws.max_row

            for row in range(start_row, end_row + 1):
                start_time = ws.cell(row, start_col - 2).value
                end_time = ws.cell(row, start_col - 1).value
                if start_time not in (None, ""):
                    current_start_time = start_time
                if end_time not in (None, ""):
                    current_end_time = end_time
                start_time = current_start_time
                end_time = current_end_time

                if not any(ws.cell(row, col["column"]).value for col in columns):
                    continue

                for col in columns:
                    value = ws.cell(row, col["column"]).value
                    if value is None or str(value).strip() == "":
                        continue
                    text = clean_schedule_name(value)
                    slot_start, slot_end, slot_hours = get_slot_schedule_values(
                        shift,
                        col["date"],
                        col["label"],
                        start_time,
                        end_time,
                    )
                    if is_schedule_placeholder(text):
                        if location == "Command Staff" or not is_schedule_opening(text):
                            continue
                        openings.append(
                            {
                                "period_label": period_label,
                                "site": site,
                                "shift": shift,
                                "location": location,
                                "opening_date": col["date"],
                                "day_name": col["day_name"],
                                "slot_label": col["label"],
                                "priority": "critical" if "CRITICAL" in text.upper() else "flex",
                                "shift_start": slot_start,
                                "shift_end": slot_end,
                            }
                        )
                        continue

                    if location == "Command Staff" or text.upper() in {"OFF", "PTO", "CLOSED"}:
                        continue

                    baseline_hours.setdefault(
                        text,
                        {"period_label": period_label, "site": site, "shift": shift, "roster_name": text, "week_1_hours": 0.0, "week_2_hours": 0.0, "pay_period_hours": 0.0},
                    )
                    baseline_hours[text][f"{week_name}_hours"] += slot_hours
                    baseline_hours[text]["pay_period_hours"] += slot_hours
                    baseline_shifts.append(
                        {
                            "period_label": period_label,
                            "site": site,
                            "shift": shift,
                            "roster_name": text,
                            "work_date": col["date"],
                            "location": location,
                        }
                    )

    return {
        "openings": openings,
        "baseline_hours": list(baseline_hours.values()),
        "baseline_shifts": baseline_shifts,
    }


def replace_overtime_period_data(cursor, period_label, site, shift, upload_id, parsed_data):
    cursor.execute(
        "DELETE FROM overtime_applications WHERE opening_id IN (SELECT id FROM overtime_openings WHERE period_label = ? AND site = ? AND shift = ?)",
        (period_label, site, shift),
    )
    cursor.execute("DELETE FROM overtime_openings WHERE period_label = ? AND site = ? AND shift = ?", (period_label, site, shift))
    cursor.execute("DELETE FROM overtime_baseline_hours WHERE period_label = ? AND site = ? AND shift = ?", (period_label, site, shift))
    cursor.execute("DELETE FROM overtime_baseline_shifts WHERE period_label = ? AND site = ? AND shift = ?", (period_label, site, shift))

    for opening in parsed_data["openings"]:
        cursor.execute(
            """
            INSERT INTO overtime_openings (
                period_label, site, shift, location, opening_date, day_name, slot_label, priority,
                shift_start, shift_end, source_upload_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                opening["period_label"],
                opening["site"],
                opening["shift"],
                opening["location"],
                opening["opening_date"],
                opening["day_name"],
                opening["slot_label"],
                opening["priority"],
                opening["shift_start"],
                opening["shift_end"],
                upload_id,
            ),
        )

    for row in parsed_data["baseline_hours"]:
        cursor.execute(
            """
            INSERT INTO overtime_baseline_hours (
                period_label, site, shift, roster_name, week_1_hours, week_2_hours, pay_period_hours, uploaded_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["period_label"],
                row["site"],
                row["shift"],
                row["roster_name"],
                row["week_1_hours"],
                row["week_2_hours"],
                row["pay_period_hours"],
                datetime.now().isoformat(),
            ),
        )

    for row in parsed_data["baseline_shifts"]:
        cursor.execute(
            """
            INSERT INTO overtime_baseline_shifts (
                period_label, site, shift, roster_name, work_date, location, uploaded_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["period_label"],
                row["site"],
                row["shift"],
                row["roster_name"],
                row["work_date"],
                row["location"],
                datetime.now().isoformat(),
            ),
            )


def remove_overtime_period_data(cursor, period_label, site, shift):
    cursor.execute(
        "SELECT file_path FROM overtime_schedule_uploads WHERE period_label = ? AND site = ? AND shift = ?",
        (period_label, site, shift),
    )
    file_paths = [row["file_path"] if isinstance(row, sqlite3.Row) else row[0] for row in cursor.fetchall()]

    cursor.execute(
        "DELETE FROM overtime_applications WHERE opening_id IN (SELECT id FROM overtime_openings WHERE period_label = ? AND site = ? AND shift = ?)",
        (period_label, site, shift),
    )
    cursor.execute("DELETE FROM overtime_openings WHERE period_label = ? AND site = ? AND shift = ?", (period_label, site, shift))
    cursor.execute("DELETE FROM overtime_baseline_hours WHERE period_label = ? AND site = ? AND shift = ?", (period_label, site, shift))
    cursor.execute("DELETE FROM overtime_baseline_shifts WHERE period_label = ? AND site = ? AND shift = ?", (period_label, site, shift))
    cursor.execute("DELETE FROM overtime_schedule_uploads WHERE period_label = ? AND site = ? AND shift = ?", (period_label, site, shift))

    for file_path in file_paths:
        try:
            if file_path and Path(file_path).exists():
                Path(file_path).unlink()
        except OSError:
            pass


def get_user_shift_for_overtime(user):
    shift_label = get_team_metadata(getattr(user, "team", "")).get("shift_label", "")
    return "day" if shift_label == "Days" else "night" if shift_label == "Nights" else ""


def get_current_overtime_period(cursor, site_scope=None):
    query = "SELECT period_label, site, shift, MAX(uploaded_at) FROM overtime_schedule_uploads"
    params = []
    if site_scope:
        query += " WHERE site = ?"
        params.append(site_scope)
    query += " GROUP BY period_label, site, shift ORDER BY MAX(uploaded_at) DESC"
    cursor.execute(query, params)
    return cursor.fetchall()


def get_user_opening_application_ids(cursor, user_id):
    cursor.execute(
        "SELECT opening_id FROM overtime_applications WHERE user_id = ? AND status = 'Applied'",
        (user_id,),
    )
    return {row[0] for row in cursor.fetchall()}


def get_overtime_period_start(cursor, period_label, site):
    cursor.execute(
        "SELECT MIN(opening_date) FROM overtime_openings WHERE period_label = ? AND site = ?",
        (period_label, site),
    )
    row = cursor.fetchone()
    if not row or not row[0]:
        return None
    return parse_iso_date(row[0])


def get_assigned_overtime_hours_map(cursor, period_label, site, period_start):
    assigned_hours = {}
    if not period_start:
        return assigned_hours

    cursor.execute(
        """
        SELECT overtime_openings.opening_date,
               overtime_openings.shift_start,
               overtime_openings.shift_end,
               COALESCE(users.first_name, '') AS first_name,
               COALESCE(users.last_name, '') AS last_name,
               users.username AS username
        FROM overtime_openings
        JOIN users ON overtime_openings.assigned_user_id = users.id
        WHERE overtime_openings.period_label = ?
          AND overtime_openings.site = ?
          AND overtime_openings.status = 'Assigned'
        """,
        (period_label, site),
    )
    for row in cursor.fetchall():
        roster_name = get_roster_name_from_values(row["first_name"], row["last_name"], row["username"])
        bucket = get_overtime_week_bucket(period_start, row["opening_date"])
        hours = assigned_hours.setdefault(roster_name, {"week_1": 0.0, "week_2": 0.0, "pay_period": 0.0})
        slot_hours = get_opening_hours_from_row(row)
        hours[bucket] += slot_hours
        hours["pay_period"] += slot_hours
    return assigned_hours


def get_overtime_dashboard_counts(cursor, user):
    site_scope = get_site_scope_for_user(user)
    query = "SELECT * FROM overtime_openings WHERE status = 'Open'"
    params = []
    if site_scope:
        query += " AND site = ?"
        params.append(site_scope)
    cursor.execute(query, params)
    openings = [enrich_overtime_opening_row(row) for row in cursor.fetchall()]
    open_count = len(filter_openings_for_user(openings, user))

    cursor.execute("SELECT COUNT(*) FROM overtime_applications WHERE user_id = ? AND status = 'Applied'", (user.id,))
    applied_count = cursor.fetchone()[0]
    return open_count, applied_count


def parse_iso_date(value):
    return datetime.strptime(value, "%Y-%m-%d").date()


def get_overtime_week_bucket(period_start, opening_date_str):
    opening_date = parse_iso_date(opening_date_str)
    return "week_1" if (opening_date - period_start).days < 7 else "week_2"


def get_opening_hours_from_row(row):
    return calculate_schedule_hours(row["shift_start"], row["shift_end"])


def auto_assign_overtime_period(cursor, period_label, site, shift, actor_user):
    cursor.execute(
        """
        SELECT *
        FROM overtime_openings
        WHERE period_label = ? AND site = ? AND shift = ? AND status = 'Open'
        ORDER BY CASE priority WHEN 'critical' THEN 0 ELSE 1 END, opening_date, location, id
        """,
        (period_label, site, shift),
    )
    openings = cursor.fetchall()
    if not openings:
        return {"assigned": 0, "unfilled": 0}

    period_start = min(parse_iso_date(row["opening_date"]) for row in openings)
    assigned_ot_hours = get_assigned_overtime_hours_map(cursor, period_label, site, period_start)

    cursor.execute(
        """
        SELECT roster_name, week_1_hours, week_2_hours, pay_period_hours
        FROM overtime_baseline_hours
        WHERE period_label = ? AND site = ? AND shift = ?
        """,
        (period_label, site, shift),
    )
    baseline_hours = {
        row["roster_name"]: {
            "week_1": float(row["week_1_hours"] or 0),
            "week_2": float(row["week_2_hours"] or 0),
            "pay_period": float(row["pay_period_hours"] or 0),
        }
        for row in cursor.fetchall()
    }

    cursor.execute(
        """
        SELECT roster_name, work_date
        FROM overtime_baseline_shifts
        WHERE period_label = ? AND site = ? AND shift = ?
        """,
        (period_label, site, shift),
    )
    assigned_dates = {}
    for row in cursor.fetchall():
        assigned_dates.setdefault(row["roster_name"], set()).add(row["work_date"])

    assigned_count = 0
    unfilled_count = 0

    for opening in openings:
        slot_hours = get_opening_hours_from_row(opening)
        bucket = get_overtime_week_bucket(period_start, opening["opening_date"])
        cursor.execute(
            """
            SELECT overtime_applications.id,
                   overtime_applications.user_id,
                   COALESCE(users.building, '') AS building,
                   COALESCE(users.first_name, '') AS first_name,
                   COALESCE(users.last_name, '') AS last_name,
                   users.username AS username
            FROM overtime_applications
            JOIN users ON overtime_applications.user_id = users.id
            WHERE overtime_applications.opening_id = ? AND overtime_applications.status = 'Applied'
            """,
            (opening["id"],),
        )
        applicants = cursor.fetchall()
        ranked = []
        for applicant in applicants:
            roster_name = get_roster_name_from_values(applicant["first_name"], applicant["last_name"], applicant["username"])
            hours = baseline_hours.setdefault(roster_name, {"week_1": 0.0, "week_2": 0.0, "pay_period": 0.0})
            ot_hours = assigned_ot_hours.setdefault(roster_name, {"week_1": 0.0, "week_2": 0.0, "pay_period": 0.0})
            taken_dates = assigned_dates.setdefault(roster_name, set())
            if opening["opening_date"] in taken_dates:
                continue
            if ot_hours[bucket] + slot_hours > MAX_OT_HOURS_PER_WEEK or ot_hours["pay_period"] + slot_hours > MAX_OT_HOURS_PER_PAY_PERIOD:
                continue
            if hours[bucket] + slot_hours > 60 or hours["pay_period"] + slot_hours > 120:
                continue
            building_match = 0 if applicant["building"] and applicant["building"] == opening["location"] else 1
            ranked.append(
                {
                    "application_id": applicant["id"],
                    "user_id": applicant["user_id"],
                    "roster_name": roster_name,
                    "score": (building_match, hours["pay_period"], hours["week_1"] + hours["week_2"], roster_name),
                }
            )

        ranked.sort(key=lambda item: item["score"])
        if not ranked:
            unfilled_count += 1
            continue

        selected = ranked[0]
        cursor.execute(
            """
            UPDATE overtime_openings
            SET status = 'Assigned', assigned_user_id = ?, assigned_at = ?, assigned_by = ?
            WHERE id = ?
            """,
            (selected["user_id"], datetime.now().isoformat(), actor_user.username, opening["id"]),
        )
        cursor.execute(
            "UPDATE overtime_applications SET status = 'Assigned' WHERE id = ?",
            (selected["application_id"],),
        )
        cursor.execute(
            "UPDATE overtime_applications SET status = 'Not Selected' WHERE opening_id = ? AND id != ? AND status = 'Applied'",
            (opening["id"], selected["application_id"]),
        )
        baseline_hours[selected["roster_name"]][bucket] += slot_hours
        assigned_ot_hours[selected["roster_name"]][bucket] += slot_hours
        assigned_ot_hours[selected["roster_name"]]["pay_period"] += slot_hours
        baseline_hours[selected["roster_name"]]["pay_period"] += slot_hours
        assigned_dates[selected["roster_name"]].add(opening["opening_date"])
        assigned_count += 1

        create_notification(
            cursor,
            selected["user_id"],
            "Overtime assigned",
            f"You were assigned overtime at {opening['location']} on {opening['opening_date']}.",
            category="success",
            link="/overtime",
        )

    log_audit(
        cursor,
        "overtime_opening",
        None,
        "auto_assigned",
        actor_user,
        details=f"Auto-assigned OT for {period_label} / {site} / {shift}: {assigned_count} assigned, {unfilled_count} unfilled",
    )
    return {"assigned": assigned_count, "unfilled": unfilled_count}


def assign_overtime_application(cursor, opening, application_id, actor_user, assignment_mode="manual"):
    cursor.execute(
        """
        SELECT overtime_applications.id,
               overtime_applications.user_id,
               overtime_applications.note,
               overtime_applications.status,
               COALESCE(users.first_name, '') AS first_name,
               COALESCE(users.last_name, '') AS last_name,
               users.username AS username
        FROM overtime_applications
        JOIN users ON overtime_applications.user_id = users.id
        WHERE overtime_applications.id = ? AND overtime_applications.opening_id = ?
        """,
        (application_id, opening["id"]),
    )
    application = cursor.fetchone()
    if not application:
        raise ValueError("That overtime application was not found.")
    if application["status"] != "Applied":
        raise ValueError("Only pending overtime applications can be assigned.")
    if opening["status"] != "Open":
        raise ValueError("That overtime opening is no longer open.")

    period_start = get_overtime_period_start(cursor, opening["period_label"], opening["site"])
    if period_start is None:
        raise ValueError("Unable to determine the overtime pay period for that opening.")

    roster_name = get_roster_name_from_values(application["first_name"], application["last_name"], application["username"])
    bucket = get_overtime_week_bucket(period_start, opening["opening_date"])
    slot_hours = get_opening_hours_from_row(opening)
    assigned_ot_hours = get_assigned_overtime_hours_map(cursor, opening["period_label"], opening["site"], period_start)
    current_ot_hours = assigned_ot_hours.setdefault(roster_name, {"week_1": 0.0, "week_2": 0.0, "pay_period": 0.0})
    if current_ot_hours[bucket] + slot_hours > MAX_OT_HOURS_PER_WEEK:
        raise ValueError(f"{get_display_name_from_values(application['first_name'], application['last_name'], application['username'])} would exceed the 20 OT hours weekly limit.")
    if current_ot_hours["pay_period"] + slot_hours > MAX_OT_HOURS_PER_PAY_PERIOD:
        raise ValueError(f"{get_display_name_from_values(application['first_name'], application['last_name'], application['username'])} would exceed the 40 OT hours pay-period limit.")

    cursor.execute(
        """
        UPDATE overtime_openings
        SET status = 'Assigned', assigned_user_id = ?, assigned_at = ?, assigned_by = ?
        WHERE id = ?
        """,
        (application["user_id"], datetime.now().isoformat(), actor_user.username, opening["id"]),
    )
    cursor.execute("UPDATE overtime_applications SET status = 'Assigned' WHERE id = ?", (application_id,))
    cursor.execute(
        "UPDATE overtime_applications SET status = 'Not Selected' WHERE opening_id = ? AND id != ? AND status = 'Applied'",
        (opening["id"], application_id),
    )

    selected_name = get_display_name_from_values(application["first_name"], application["last_name"], application["username"])
    create_notification(
        cursor,
        application["user_id"],
        "Overtime assigned",
        f"You were assigned overtime at {opening['location']} on {opening['opening_date']}.",
        category="success",
        link="/overtime",
    )
    log_audit(
        cursor,
        "overtime_opening",
        opening["id"],
        "manually_assigned" if assignment_mode == "manual" else "assigned",
        actor_user,
        target_user_id=application["user_id"],
        details=f"{selected_name} assigned to OT opening {opening['location']} on {opening['opening_date']}",
    )

    return {
        "user_id": application["user_id"],
        "display_name": selected_name,
    }


def revoke_overtime_assignment(cursor, opening, actor_user):
    if opening["status"] != "Assigned" or not opening["assigned_user_id"]:
        raise ValueError("That overtime opening is not currently assigned.")

    cursor.execute(
        """
        SELECT overtime_applications.id,
               COALESCE(users.first_name, '') AS first_name,
               COALESCE(users.last_name, '') AS last_name,
               users.username AS username
        FROM overtime_applications
        JOIN users ON overtime_applications.user_id = users.id
        WHERE overtime_applications.opening_id = ? AND overtime_applications.user_id = ? AND overtime_applications.status = 'Assigned'
        """,
        (opening["id"], opening["assigned_user_id"]),
    )
    assigned_application = cursor.fetchone()

    cursor.execute(
        """
        UPDATE overtime_openings
        SET status = 'Open', assigned_user_id = NULL, assigned_at = NULL, assigned_by = NULL
        WHERE id = ?
        """,
        (opening["id"],),
    )
    if assigned_application:
        cursor.execute(
            "UPDATE overtime_applications SET status = 'Revoked' WHERE id = ?",
            (assigned_application["id"],),
        )
        assigned_name = get_display_name_from_values(
            assigned_application["first_name"],
            assigned_application["last_name"],
            assigned_application["username"],
        )
    else:
        assigned_name = "the assigned officer"

    create_notification(
        cursor,
        opening["assigned_user_id"],
        "Overtime revoked",
        f"Your overtime assignment for {opening['location']} on {opening['opening_date']} was revoked.",
        category="warning",
        link="/overtime",
    )
    log_audit(
        cursor,
        "overtime_opening",
        opening["id"],
        "revoked",
        actor_user,
        target_user_id=opening["assigned_user_id"],
        details=f"Revoked OT assignment for {assigned_name} at {opening['location']} on {opening['opening_date']}",
    )


def build_overtime_admin_applicant_details(cursor, openings):
    if not openings:
        return {}

    opening_ids = [row["id"] for row in openings]
    placeholders = ",".join("?" for _ in opening_ids)

    baseline_keys = {(row["period_label"], row["site"], row["shift"]) for row in openings}
    assigned_hours_map = {}

    for period_label, site, shift in baseline_keys:
        period_start = get_overtime_period_start(cursor, period_label, site)
        ot_hours = get_assigned_overtime_hours_map(cursor, period_label, site, period_start)
        for roster_name, hours in ot_hours.items():
            assigned_hours_map[(period_label, site, shift, roster_name)] = round(hours["pay_period"], 2)

    cursor.execute(
        f"""
        SELECT overtime_applications.opening_id,
               overtime_applications.id AS application_id,
               overtime_applications.user_id,
               overtime_applications.note,
               overtime_applications.status,
               overtime_applications.submitted_at,
               COALESCE(users.first_name, '') AS first_name,
               COALESCE(users.last_name, '') AS last_name,
               users.username,
               COALESCE(users.building, '') AS building
        FROM overtime_applications
        JOIN users ON overtime_applications.user_id = users.id
        WHERE overtime_applications.opening_id IN ({placeholders})
          AND overtime_applications.status IN ('Applied', 'Assigned')
        ORDER BY overtime_applications.submitted_at ASC
        """,
        opening_ids,
    )

    opening_lookup = {row["id"]: row for row in openings}
    details_map = {opening_id: [] for opening_id in opening_ids}

    for row in cursor.fetchall():
        opening = opening_lookup[row["opening_id"]]
        roster_name = get_roster_name_from_values(row["first_name"], row["last_name"], row["username"])
        base_key = (opening["period_label"], opening["site"], opening["shift"], roster_name)
        current_hours = assigned_hours_map.get(base_key, 0.0)
        details_map[row["opening_id"]].append(
            {
                "application_id": row["application_id"],
                "user_id": row["user_id"],
                "display_name": get_display_name_from_values(row["first_name"], row["last_name"], row["username"]),
                "building": row["building"] or "Unassigned",
                "current_hours": round(current_hours, 2),
                "note": row["note"] or "",
                "status": row["status"],
                "submitted_at": row["submitted_at"],
            }
        )

    return details_map


def save_pto_document(file_storage, user_id):
    filename = secure_filename(file_storage.filename or "")
    if not filename:
        raise ValueError("Documentation file is missing a filename.")
    if not is_allowed_document(filename):
        raise ValueError("Documentation must be a PNG, JPG, JPEG, or WEBP image.")

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    saved_name = f"user_{user_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
    destination = UPLOAD_DIR / saved_name
    file_storage.save(destination)
    return str(destination)


def get_active_time_segment(entry):
    if not entry:
        return None
    if entry["break1_start"] and not entry["break1_end"]:
        return "break_1"
    if entry["break2_start"] and not entry["break2_end"]:
        return "break_2"
    if entry["lunch_start"] and not entry["lunch_end"]:
        return "lunch"
    return None


def get_time_entry_summary(entry, reference_iso=None):
    if not entry:
        return {
            "gross_hours": None,
            "worked_hours": None,
            "break_1_minutes": 0,
            "break_2_minutes": 0,
            "unpaid_break_minutes": 0,
            "paid_lunch_minutes": 0,
            "allowed_breaks": 0,
            "active_segment": None,
            "can_take_second_break": False,
            "breaks_taken": 0,
            "lunch_taken": False,
        }

    active_segment = get_active_time_segment(entry)
    break_1_minutes = calculate_duration_minutes(
        entry["break1_start"],
        entry["break1_end"],
        reference_iso if active_segment == "break_1" else None,
    )
    break_2_minutes = calculate_duration_minutes(
        entry["break2_start"],
        entry["break2_end"],
        reference_iso if active_segment == "break_2" else None,
    )
    paid_lunch_minutes = calculate_duration_minutes(
        entry["lunch_start"],
        entry["lunch_end"],
        reference_iso if active_segment == "lunch" else None,
    )
    unpaid_break_minutes = round(break_1_minutes + break_2_minutes, 2)
    allowed_breaks = get_allowed_unpaid_breaks(entry["expected_hours"])
    clock_out_value = entry["clock_out"] or reference_iso
    gross_hours = calculate_gross_hours(entry["clock_in"], clock_out_value) if clock_out_value else None
    worked_hours = calculate_worked_hours(entry["clock_in"], clock_out_value, unpaid_break_minutes) if clock_out_value else None
    completed_breaks = int(bool(entry["break1_end"])) + int(bool(entry["break2_end"]))

    return {
        "gross_hours": gross_hours,
        "worked_hours": worked_hours,
        "break_1_minutes": break_1_minutes,
        "break_2_minutes": break_2_minutes,
        "unpaid_break_minutes": unpaid_break_minutes,
        "paid_lunch_minutes": paid_lunch_minutes,
        "allowed_breaks": allowed_breaks,
        "active_segment": active_segment,
        "can_take_second_break": allowed_breaks >= 2 and bool(entry["break1_end"]),
        "breaks_taken": completed_breaks,
        "lunch_taken": bool(entry["lunch_end"]),
    }


def refresh_time_clock_totals(cursor, entry_id):
    cursor.execute(
        """
        SELECT *
        FROM time_clock_entries
        WHERE id = ?
        """,
        (entry_id,),
    )
    entry = cursor.fetchone()
    if not entry:
        return None
    summary = get_time_entry_summary(entry)
    worked_hours = summary["worked_hours"] if entry["clock_out"] else None
    cursor.execute(
        """
        UPDATE time_clock_entries
        SET worked_hours = ?, unpaid_break_minutes = ?, paid_lunch_minutes = ?, updated_at = ?
        WHERE id = ?
        """,
        (
            worked_hours,
            summary["unpaid_break_minutes"],
            summary["paid_lunch_minutes"],
            datetime.now().isoformat(),
            entry_id,
        ),
    )
    return summary


def get_time_clock_action_label(action):
    labels = {
        "clock_in": "clocked in",
        "clock_out": "clocked out",
        "start_break_1": "started unpaid break 1",
        "end_break_1": "ended unpaid break 1",
        "start_break_2": "started unpaid break 2",
        "end_break_2": "ended unpaid break 2",
        "start_lunch": "started paid lunch",
        "end_lunch": "ended paid lunch",
    }
    return labels.get(action, action.replace("_", " "))


def get_app_base_url():
    return os.getenv("PTO_APP_BASE_URL", "http://127.0.0.1:5000").rstrip("/")


def user_can_view_all_sites(user):
    return getattr(user, "role", "") == "admin"


def get_site_scope_for_user(user):
    if user_can_view_all_sites(user):
        return None
    site = (getattr(user, "site", "") or "").strip()
    return site or None


def user_has_site_access(user, site_name):
    site_name = (site_name or "").strip()
    scope = get_site_scope_for_user(user)
    return scope is None or scope == site_name


def get_command_staff_users(cursor, site_scope=None):
    if site_scope:
        cursor.execute(
            """
            SELECT id, username, email, role, rank, site
            FROM users
            WHERE email != ''
              AND (
                    role = 'admin'
                    OR (rank IN ('Captain', 'Director') AND site = ?)
              )
            ORDER BY username
            """,
            (site_scope,),
        )
    else:
        cursor.execute(
            """
            SELECT id, username, email, role, rank, site
            FROM users
            WHERE email != ''
              AND (role = 'admin' OR rank IN ('Captain', 'Director'))
            ORDER BY username
            """
        )
    return cursor.fetchall()


def get_command_staff_emails(cursor, site_scope=None):
    return [row[2] for row in get_command_staff_users(cursor, site_scope) if row[2]]


def get_broadcast_recipients(cursor, site_scope=None, rank_filter="", team_filter=""):
    query = """
        SELECT id, username, email, rank, team, site,
               COALESCE(first_name, '') AS first_name,
               COALESCE(last_name, '') AS last_name
        FROM users
        WHERE email != ''
    """
    params = []
    if site_scope:
        query += " AND site = ?"
        params.append(site_scope)
    if rank_filter:
        query += " AND rank = ?"
        params.append(rank_filter)
    if team_filter:
        query += " AND team = ?"
        params.append(team_filter)
    query += " ORDER BY last_name, first_name, username"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    recipients = []
    for row in rows:
        recipients.append(
            {
                "id": row[0],
                "username": row[1],
                "email": row[2],
                "rank": row[3] or "Officer",
                "team": row[4] or "",
                "site": row[5] or "",
                "display_name": get_display_name_from_values(row[6], row[7], row[1]),
            }
        )
    return recipients


def parse_selected_user_ids(values):
    selected_ids = []
    for value in values:
        text = str(value).strip()
        if text.isdigit():
            selected_ids.append(int(text))
    return selected_ids




def create_notification(cursor, user_id, title, message, category="info", link=""):
    cursor.execute(
        """
        INSERT INTO notifications (user_id, title, message, category, link, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (user_id, title, message, category, link, datetime.now().isoformat()),
    )


def create_notifications_for_users(cursor, user_ids, title, message, category="info", link=""):
    for user_id in user_ids:
        create_notification(cursor, user_id, title, message, category, link)


def log_audit(cursor, entity_type, entity_id, action, actor_user, target_user_id=None, details=""):
    cursor.execute(
        """
        INSERT INTO audit_log (
            entity_type, entity_id, action, actor_user_id, actor_username, target_user_id, details, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            entity_type,
            entity_id,
            action,
            getattr(actor_user, "id", None),
            getattr(actor_user, "username", ""),
            target_user_id,
            details,
            datetime.now().isoformat(),
        ),
    )


def send_notification_email(subject, body, recipients):
    if not recipients:
        return "No command staff email addresses are on file."

    smtp_host = os.getenv("PTO_SMTP_HOST", "").strip()
    smtp_port = int(os.getenv("PTO_SMTP_PORT", "587"))
    smtp_username = os.getenv("PTO_SMTP_USERNAME", "").strip()
    smtp_password = os.getenv("PTO_SMTP_PASSWORD", "").strip()
    smtp_from = os.getenv("PTO_SMTP_FROM", smtp_username).strip()
    use_tls = os.getenv("PTO_SMTP_USE_TLS", "true").lower() in {"1", "true", "yes", "on"}

    if not smtp_host or not smtp_from:
        return "Email notifications are not configured yet."

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = smtp_from
    message["To"] = ", ".join(recipients)
    message.set_content(body)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            if use_tls:
                server.starttls()
            if smtp_username and smtp_password:
                server.login(smtp_username, smtp_password)
            server.send_message(message)
    except Exception as exc:
        return f"Email notification failed: {exc}"

    return None



def user_has_rank_access(user, allowed_ranks):
    return getattr(user, "role", "") == "admin" or getattr(user, "rank", "") in allowed_ranks


def get_user_record(cursor, user_id):
    cursor.execute(
        """
        SELECT id, username, role, rank, email, phone, team, site, COALESCE(building, ''), COALESCE(first_name, ''), COALESCE(last_name, '')
        FROM users
        WHERE id = ?
        """,
        (user_id,),
    )
    row = cursor.fetchone()
    if not row:
        return None
    return {
        "id": row[0],
        "username": row[1],
        "role": row[2],
        "rank": row[3] or "Officer",
        "email": row[4] or "",
        "phone": row[5] or "",
        "team": row[6] or "",
        "site": row[7] or "",
        "building": row[8] or "",
        "first_name": row[9] or "",
        "last_name": row[10] or "",
        "display_name": get_display_name_from_values(row[9], row[10], row[1]),
    }


def count_approved_team_time_off(cursor, team, day_str, exclude_request_id=None):
    if not team:
        return 0

    query = """
        SELECT COUNT(*)
        FROM pto_requests
        JOIN users ON pto_requests.user_id = users.id
        WHERE pto_requests.status = 'Approved'
          AND pto_requests.start_date <= ?
          AND pto_requests.end_date >= ?
          AND COALESCE(users.team, '') = ?
    """
    params = [day_str, day_str, team]
    if exclude_request_id is not None:
        query += " AND pto_requests.id != ?"
        params.append(exclude_request_id)
    cursor.execute(query, params)
    return cursor.fetchone()[0]


def send_user_email(cursor, user_id, subject, body):
    user = get_user_record(cursor, user_id)
    if not user or not user["email"]:
        return None
    return send_notification_email(subject, body, [user["email"]])


def parse_datetime_local(value):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(value, fmt).isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value).isoformat()
    except ValueError:
        return None


def format_datetime_for_input(value):
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value).strftime("%Y-%m-%dT%H:%M")
    except ValueError:
        return ""


def format_datetime_display(value):
    if not value:
        return "-"
    try:
        return datetime.fromisoformat(value).strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return value.replace("T", " ")


def get_schedule_start_datetime(team, clock_date_str, schedule_note=""):
    if not team:
        return None
    team_meta = get_team_metadata(team)
    start_hour = 6 if team_meta["shift_label"] == "Days" else 18
    if "second half" in (schedule_note or "").lower():
        start_hour = 18
    elif "first half" in (schedule_note or "").lower():
        start_hour = 6
    try:
        clock_date = datetime.strptime(clock_date_str, "%Y-%m-%d")
    except ValueError:
        return None
    return clock_date.replace(hour=start_hour, minute=0, second=0, microsecond=0)


def late_punch_minutes(clock_in_str, team, clock_date_str, schedule_note=""):
    if not clock_in_str:
        return None
    schedule_start = get_schedule_start_datetime(team, clock_date_str, schedule_note)
    if not schedule_start:
        return None
    try:
        clock_in = datetime.fromisoformat(clock_in_str)
    except ValueError:
        return None
    delta_minutes = int((clock_in - schedule_start).total_seconds() / 60)
    if delta_minutes >= LATE_PUNCH_THRESHOLD_MINUTES:
        return delta_minutes
    return None


def build_timecard_review_context(start_date, end_date, selected_user_id="", site_scope=None):
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    user_query = "SELECT id, username, rank, team, site FROM users"
    user_params = []
    if site_scope:
        user_query += " WHERE site = ?"
        user_params.append(site_scope)
    user_query += " ORDER BY username"
    c.execute(user_query, user_params)
    users = [dict(row) for row in c.fetchall()]

    query = """
        SELECT
            time_clock_entries.*,
            users.username,
            users.rank,
            users.site,
            COALESCE(time_clock_entries.team, users.team, '') AS resolved_team,
            COALESCE(time_clock_entries.schedule_note, '') AS resolved_schedule_note
        FROM time_clock_entries
        JOIN users ON time_clock_entries.user_id = users.id
        WHERE time_clock_entries.clock_date BETWEEN ? AND ?
    """
    params = [start_date, end_date]
    if site_scope:
        query += " AND users.site = ?"
        params.append(site_scope)

    if selected_user_id:
        query += " AND time_clock_entries.user_id = ?"
        params.append(selected_user_id)
    query += " ORDER BY time_clock_entries.clock_date DESC, users.username"
    c.execute(query, params)
    raw_entries = [dict(row) for row in c.fetchall()]

    entries = []
    entry_map = {}
    for entry in raw_entries:
        entry["team"] = entry["resolved_team"] or ""
        entry["schedule_note"] = entry["resolved_schedule_note"] or ""
        entry["team_meta"] = get_team_metadata(entry["team"])
        entry_summary = get_time_entry_summary(entry)
        entry.update(entry_summary)
        entry["variance"] = None
        if entry["worked_hours"] is not None and entry["expected_hours"] is not None:
            entry["variance"] = round(entry["worked_hours"] - entry["expected_hours"], 2)
        entry["clock_in_display"] = format_datetime_display(entry["clock_in"])
        entry["clock_out_display"] = format_datetime_display(entry["clock_out"])
        entry["corrected_at_display"] = format_datetime_display(entry["corrected_at"])
        entry["clock_in_input"] = format_datetime_for_input(entry["clock_in"])
        entry["clock_out_input"] = format_datetime_for_input(entry["clock_out"])
        entry["late_minutes"] = late_punch_minutes(entry["clock_in"], entry["team"], entry["clock_date"], entry["schedule_note"])
        entries.append(entry)
        entry_map[(entry["user_id"], entry["clock_date"])] = entry

    scope_users = [user for user in users if not selected_user_id or str(user["id"]) == str(selected_user_id)]
    date_strings = [date.strftime("%Y-%m-%d") for date in iter_dates(start_date, end_date)]

    employee_summaries = []
    exceptions = []
    variance_threshold = 0.25

    for user in scope_users:
        team = user["team"] or ""
        user_site = user.get("site", "") or ""
        expected_total = 0.0
        worked_total = 0.0
        unpaid_break_total = 0.0
        paid_lunch_total = 0.0
        missing_punches = 0
        unscheduled_punches = 0
        open_shifts = 0
        late_punches = 0

        for date_str in date_strings:
            schedule = get_team_schedule(team, date_str)
            entry = entry_map.get((user["id"], date_str))

            if schedule["expected_hours"] is not None:
                expected_total += schedule["expected_hours"]

            if entry and entry["worked_hours"] is not None:
                worked_total += entry["worked_hours"]
                unpaid_break_total += entry["unpaid_break_minutes"]
                paid_lunch_total += entry["paid_lunch_minutes"]

            if schedule["expected_hours"] and not entry:
                missing_punches += 1
                exceptions.append(
                    {
                        "type": "Missing punch",
                        "severity": "critical",
                        "username": user["username"],
                        "team": team or "",
                        "date": date_str,
                        "details": f"Scheduled for {schedule['expected_hours']:.2f} hours with no time entry.",
                    }
                )

            if entry and entry["clock_in"] and not entry["clock_out"]:
                open_shifts += 1
                active_segment = entry["active_segment"]
                details = "Clocked in without a matching clock-out."
                if active_segment == "break_1":
                    details = "Employee is still out on unpaid break 1."
                elif active_segment == "break_2":
                    details = "Employee is still out on unpaid break 2."
                elif active_segment == "lunch":
                    details = "Employee is still out on paid lunch."
                exceptions.append(
                    {
                        "type": "Open shift",
                        "severity": "warning",
                        "username": user["username"],
                        "team": team or "",
                        "date": date_str,
                        "details": details,
                    }
                )

            if entry and (entry["clock_in"] or entry["worked_hours"] is not None) and (entry["expected_hours"] or 0) == 0:
                unscheduled_punches += 1
                exceptions.append(
                    {
                        "type": "Unscheduled punch",
                        "severity": "warning",
                        "username": user["username"],
                        "team": team or "",
                        "date": date_str,
                        "details": f"Time was recorded on a scheduled off day ({schedule['schedule_note']}).",
                    }
                )

            late_minutes = entry["late_minutes"] if entry else None
            if late_minutes is not None and (entry["expected_hours"] or 0) > 0:
                late_punches += 1
                exceptions.append(
                    {
                        "type": "Late punch",
                        "severity": "warning",
                        "username": user["username"],
                        "team": team or "",
                        "date": date_str,
                        "details": f"Clock-in was {late_minutes} minutes after the scheduled start.",
                    }
                )

            if entry and entry["worked_hours"] is not None and entry["expected_hours"] is not None and abs(entry["worked_hours"] - entry["expected_hours"]) >= variance_threshold:
                exceptions.append(
                    {
                        "type": "Variance",
                        "severity": "info",
                        "username": user["username"],
                        "team": team or "",
                        "date": date_str,
                        "details": f"Variance of {entry['worked_hours'] - entry['expected_hours']:+.2f} hours. Unpaid breaks tracked: {entry['unpaid_break_minutes']:.0f} minutes.",
                    }
                )

        employee_summaries.append(
            {
                "user_id": user["id"],
                "username": user["username"],
                "rank": user["rank"] or "Officer",
                "team": team or "",
                "team_meta": get_team_metadata(team),
                "expected_total": round(expected_total, 2),
                "worked_total": round(worked_total, 2),
                "variance": round(worked_total - expected_total, 2),
                "unpaid_break_minutes_total": round(unpaid_break_total, 2),
                "paid_lunch_minutes_total": round(paid_lunch_total, 2),
                "missing_punches": missing_punches,
                "open_shifts": open_shifts,
                "unscheduled_punches": unscheduled_punches,
                "late_punches": late_punches,
                "exception_count": missing_punches + open_shifts + unscheduled_punches + late_punches,
            }
        )

    employee_summaries.sort(key=lambda item: (item["team"], item["username"]))
    exceptions.sort(key=lambda item: (item["date"], item["username"], item["type"]))
    conn.close()

    return {
        "users": users,
        "entries": entries,
        "employee_summaries": employee_summaries,
        "exceptions": exceptions,
        "variance_threshold": variance_threshold,
    }


def get_pay_period_bounds(reference_date=None):
    anchor_date = datetime.strptime(ROTATION_START_DATE, "%Y-%m-%d").date()
    reference = reference_date or datetime.now().date()
    day_offset = (reference - anchor_date).days
    period_offset = (day_offset // 14) * 14
    period_start = anchor_date + timedelta(days=period_offset)
    period_end = period_start + timedelta(days=13)
    return period_start.strftime("%Y-%m-%d"), period_end.strftime("%Y-%m-%d")


def build_command_site_summaries(user):
    if not user_has_rank_access(user, COMMAND_REVIEW_RANKS):
        return []

    scope = get_site_scope_for_user(user)
    visible_sites = [scope] if scope else list(SITE_OPTIONS)
    summaries = []
    today_str = datetime.now().strftime("%Y-%m-%d")

    for site_name in visible_sites:
        conn = sqlite3.connect(DB)
        c = conn.cursor()

        c.execute(
            """
            SELECT COUNT(*)
            FROM pto_requests
            JOIN users ON pto_requests.user_id = users.id
            WHERE pto_requests.status = 'Pending' AND users.site = ?
            """,
            (site_name,),
        )
        pending_pto = c.fetchone()[0]

        c.execute(
            """
            SELECT COUNT(*)
            FROM pto_requests
            JOIN users ON pto_requests.user_id = users.id
            WHERE pto_requests.status = 'Approved'
              AND users.site = ?
              AND pto_requests.start_date <= ?
              AND pto_requests.end_date >= ?
            """,
            (site_name, today_str, today_str),
        )
        approved_today = c.fetchone()[0]

        c.execute(
            """
            SELECT COUNT(*)
            FROM pto_requests
            JOIN users ON pto_requests.user_id = users.id
            WHERE pto_requests.status = 'Pending'
              AND users.site = ?
              AND COALESCE(pto_requests.documentation_path, '') != ''
            """,
            (site_name,),
        )
        pending_documents = c.fetchone()[0]

        c.execute(
            """
            SELECT COUNT(*)
            FROM shift_swaps
            JOIN users AS requester ON shift_swaps.requesting_user_id = requester.id
            WHERE shift_swaps.status = 'Pending' AND requester.site = ?
            """,
            (site_name,),
        )
        pending_swaps = c.fetchone()[0]

        c.execute(
            """
            SELECT COUNT(*)
            FROM notifications
            JOIN users ON notifications.user_id = users.id
            WHERE notifications.is_read = 0 AND users.site = ?
            """,
            (site_name,),
        )
        unread_alerts = c.fetchone()[0]
        conn.close()

        summaries.append(
            {
                "site": site_name,
                "pending_pto": pending_pto,
                "pending_swaps": pending_swaps,
                "approved_today": approved_today,
                "pending_documents": pending_documents,
                "unread_alerts": unread_alerts,
            }
        )

    return summaries


def iter_dates(start_date_str, end_date_str):
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    if end_date < start_date:
        start_date, end_date = end_date, start_date
    current = start_date
    while current <= end_date:
        yield current
        current += timedelta(days=1)


# ---------------- USER CLASS ----------------
class User(UserMixin):
    def __init__(self, id_, username, password, role, rank="Officer", email="", phone="", team="", site="", building="", first_name="", last_name=""):
        self.id = id_
        self.username = username
        self.password = password
        self.role = role
        self.rank = rank
        self.email = email
        self.phone = phone
        self.team = team
        self.site = site
        self.building = building or ""
        self.first_name = first_name or ""
        self.last_name = last_name or ""

    @property
    def display_name(self):
        return get_display_name_from_values(self.first_name, self.last_name, self.username)

@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute(
        """
        SELECT id, username, password, role, rank, email, phone, team, site, COALESCE(building, ''), COALESCE(first_name, ''), COALESCE(last_name, '')
        FROM users
        WHERE id=?
        """,
        (user_id,),
    )
    user = c.fetchone()
    conn.close()
    if user:
        return User(*user)
    return None

# ---------------- LOGIN ----------------
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = sqlite3.connect(DB)
        c = conn.cursor()
        c.execute(
            """
            SELECT id, username, password, role, rank, email, phone, team, site, COALESCE(building, ''), COALESCE(first_name, ''), COALESCE(last_name, '')
            FROM users
            WHERE username=?
            """,
            (username,),
        )
        user = c.fetchone()
        conn.close()

        if user and check_password_hash(user[2], password):
            login_user(User(*user))
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid login")

    return render_template("login.html")

# ---------------- REGISTER ----------------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        first_name = request.form["first_name"].strip()
        last_name = request.form["last_name"].strip()
        username = request.form["username"].strip()
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]
        email = request.form["email"].strip()
        phone = normalize_phone_number(request.form["phone"])

        if not first_name or not last_name:
            flash("First and last name are required.")
            return redirect(url_for("register"))

        if password != confirm_password:
            flash("Passwords do not match.")
            return redirect(url_for("register"))

        conn = sqlite3.connect(DB)
        c = conn.cursor()

        try:
            c.execute(
                "INSERT INTO users (username, password, role, rank, email, phone, team, site, first_name, last_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (username, generate_password_hash(password), "officer", "Officer", email, phone, "", "", first_name, last_name)
            )
            conn.commit()
            flash("Account created successfully.")
            return redirect(url_for("login"))
        except sqlite3.IntegrityError:
            flash("That username already exists.")
        finally:
            conn.close()

    return render_template("register.html")



# ---------------- DASHBOARD ----------------
@app.route("/dashboard")
@login_required
def dashboard():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute("SELECT * FROM pto_requests WHERE user_id=? ORDER BY created_at DESC", (current_user.id,))
    requests = c.fetchall()

    c.execute(
        """
        SELECT
            shift_swaps.id,
            shift_swaps.requesting_user_id,
            shift_swaps.swap_with_user_id,
            shift_swaps.start_date,
            shift_swaps.end_date,
            shift_swaps.reason,
            shift_swaps.status,
            shift_swaps.approved_by,
            shift_swaps.created_at,
            COALESCE(shift_swaps.review_note, ''),
            users.username
        FROM shift_swaps
        LEFT JOIN users ON shift_swaps.swap_with_user_id = users.id
        WHERE shift_swaps.requesting_user_id = ?
        ORDER BY shift_swaps.created_at DESC
        """,
        (current_user.id,),
    )
    shift_swaps = c.fetchall()

    today = datetime.now().strftime("%Y-%m-%d")
    today_schedule = get_team_schedule(current_user.team, today)
    c.execute(
        """
        SELECT *
        FROM time_clock_entries
        WHERE user_id = ? AND clock_date = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (current_user.id, today),
    )
    today_time_entry = c.fetchone()

    c.execute(
        """
        SELECT *
        FROM time_clock_entries
        WHERE user_id = ?
        ORDER BY clock_date DESC, id DESC
        LIMIT 5
        """,
        (current_user.id,),
    )
    recent_time_entries = c.fetchall()

    c.execute(
        """
        SELECT id, title, message, category, link, is_read, created_at
        FROM notifications
        WHERE user_id = ?
        ORDER BY created_at DESC
        LIMIT 6
        """,
        (current_user.id,),
    )
    notifications = c.fetchall()

    c.execute("SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0", (current_user.id,))
    unread_notifications = c.fetchone()[0]

    overtime_open_count, my_overtime_app_count = get_overtime_dashboard_counts(c, current_user)

    pay_period_start, pay_period_end = get_pay_period_bounds()
    c.execute(
        """
        SELECT COALESCE(SUM(worked_hours), 0), COALESCE(SUM(expected_hours), 0)
        FROM time_clock_entries
        WHERE user_id = ? AND clock_date BETWEEN ? AND ?
        """,
        (current_user.id, pay_period_start, pay_period_end),
    )
    pay_period_worked, pay_period_expected = c.fetchone()

    conn.close()

    command_site_summaries = build_command_site_summaries(current_user)
    command_summary_totals = {
        "pending_pto": sum(item["pending_pto"] for item in command_site_summaries),
        "pending_swaps": sum(item["pending_swaps"] for item in command_site_summaries),
        "approved_today": sum(item["approved_today"] for item in command_site_summaries),
        "pending_documents": sum(item["pending_documents"] for item in command_site_summaries),
        "unread_alerts": sum(item["unread_alerts"] for item in command_site_summaries),
    }

    team_meta = get_team_metadata(current_user.team)
    pto_pending = len([request for request in requests if request[4] == "Pending"])
    swap_pending = len([swap for swap in shift_swaps if swap[6] == "Pending"])

    return render_template(
        "dashboard.html",
        requests=requests,
        shift_swaps=shift_swaps,
        today_time_entry=today_time_entry,
        today_schedule=today_schedule,
        team_meta=team_meta,
        current_site=current_user.site,
        recent_time_entries=recent_time_entries,
        notifications=notifications,
        unread_notifications=unread_notifications,
        overtime_open_count=overtime_open_count,
        my_overtime_app_count=my_overtime_app_count,
        pto_pending=pto_pending,
        swap_pending=swap_pending,
        pay_period_start=pay_period_start,
        pay_period_end=pay_period_end,
        pay_period_worked=pay_period_worked or 0,
        pay_period_expected=pay_period_expected or 0,
        command_site_summaries=command_site_summaries,
        command_summary_totals=command_summary_totals,
        is_command_staff=user_has_rank_access(current_user, COMMAND_REVIEW_RANKS),
    )


@app.route("/notifications", methods=["POST"])
@login_required
def notifications_action():
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    action = request.form.get("action", "").strip().lower()
    notification_id = request.form.get("notification_id", "").strip()

    if action == "mark_read" and notification_id:
        c.execute(
            """
            UPDATE notifications
            SET is_read = 1
            WHERE id = ? AND user_id = ?
            """,
            (notification_id, current_user.id),
        )
        conn.commit()
        flash("Notification marked as read.")
    elif action == "mark_all_read":
        c.execute("UPDATE notifications SET is_read = 1 WHERE user_id = ?", (current_user.id,))
        conn.commit()
        flash("All notifications marked as read.")
    elif action == "clear_read":
        c.execute("DELETE FROM notifications WHERE user_id = ? AND is_read = 1", (current_user.id,))
        conn.commit()
        flash("Read notifications cleared.")
    else:
        flash("Unknown notification action.")

    conn.close()
    return redirect(url_for("dashboard"))


@app.route("/overtime", methods=["GET", "POST"])
@login_required
def overtime():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    site_scope = get_site_scope_for_user(current_user)
    allowed_teams = get_opposite_ot_teams_for_user(getattr(current_user, "team", "") or "")

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        opening_id = request.form.get("opening_id", "").strip()

        if not opening_id:
            conn.close()
            flash("Please choose a valid overtime opening.")
            return redirect(url_for("overtime"))

        c.execute("SELECT * FROM overtime_openings WHERE id = ?", (opening_id,))
        opening = c.fetchone()
        if not opening:
            conn.close()
            flash("Overtime opening not found.")
            return redirect(url_for("overtime"))
        if site_scope and opening["site"] != site_scope:
            conn.close()
            flash("You can only apply to openings for your site.")
            return redirect(url_for("overtime"))
        opening_view = enrich_overtime_opening_row(opening)
        if not user_is_command_team(current_user) and allowed_teams and opening_view["team"] not in allowed_teams:
            conn.close()
            flash("You can only apply to overtime for the opposite team rotation.")
            return redirect(url_for("overtime"))
        if opening["status"] != "Open":
            conn.close()
            flash("That overtime opening is no longer open.")
            return redirect(url_for("overtime"))

        if action == "apply":
            c.execute(
                "SELECT 1 FROM overtime_applications WHERE opening_id = ? AND user_id = ? AND status IN ('Applied', 'Assigned')",
                (opening_id, current_user.id),
            )
            if c.fetchone():
                flash("You have already applied for that overtime opening.")
            else:
                note = request.form.get("note", "").strip()
                c.execute(
                    """
                    INSERT INTO overtime_applications (opening_id, user_id, note, submitted_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    (opening_id, current_user.id, note, datetime.now().isoformat()),
                )
                create_notification(
                    c,
                    current_user.id,
                    "Overtime application submitted",
                    f"You applied for overtime at {opening['location']} on {opening['opening_date']}.",
                    category="info",
                    link="/overtime",
                )
                log_audit(
                    c,
                    "overtime_application",
                    c.lastrowid,
                    "submitted",
                    current_user,
                    details=f"Applied for OT opening {opening['location']} on {opening['opening_date']}",
                )
                conn.commit()
                flash("Overtime application submitted.")

        elif action == "withdraw":
            c.execute(
                """
                UPDATE overtime_applications
                SET status = 'Withdrawn'
                WHERE opening_id = ? AND user_id = ? AND status = 'Applied'
                """,
                (opening_id, current_user.id),
            )
            if c.rowcount:
                log_audit(
                    c,
                    "overtime_application",
                    None,
                    "withdrawn",
                    current_user,
                    details=f"Withdrew OT application for {opening['location']} on {opening['opening_date']}",
                )
                conn.commit()
                flash("Overtime application withdrawn.")
            else:
                flash("No active application was found for that opening.")

        conn.close()
        return redirect(url_for("overtime"))

    openings_query = """
        SELECT overtime_openings.*,
               COALESCE(u.first_name, '') AS assigned_first_name,
               COALESCE(u.last_name, '') AS assigned_last_name,
               u.username AS assigned_username,
               (
                 SELECT COUNT(*)
                 FROM overtime_applications oa
                 WHERE oa.opening_id = overtime_openings.id AND oa.status = 'Applied'
               ) AS applicant_count
        FROM overtime_openings
        LEFT JOIN users u ON overtime_openings.assigned_user_id = u.id
        WHERE 1 = 1
    """
    params = []
    if site_scope:
        openings_query += " AND overtime_openings.site = ?"
        params.append(site_scope)
    openings_query += " ORDER BY overtime_openings.opening_date, CASE overtime_openings.priority WHEN 'critical' THEN 0 ELSE 1 END, overtime_openings.location"
    c.execute(openings_query, params)
    openings = [enrich_overtime_opening_row(row) for row in c.fetchall()]
    openings = filter_openings_for_user(openings, current_user)
    openings = sorted(
        openings,
        key=lambda opening: (
            opening["opening_date"],
            0 if str(opening["priority"]).lower() == "critical" else 1,
            opening["location"],
            opening["shift_start"],
        ),
    )

    applied_ids = get_user_opening_application_ids(c, current_user.id)
    c.execute(
        """
        SELECT overtime_applications.*, overtime_openings.location, overtime_openings.opening_date, overtime_openings.priority, overtime_openings.shift, overtime_openings.shift_start, overtime_openings.slot_label
        FROM overtime_applications
        JOIN overtime_openings ON overtime_applications.opening_id = overtime_openings.id
        WHERE overtime_applications.user_id = ?
        ORDER BY overtime_applications.submitted_at DESC
        LIMIT 12
        """,
        (current_user.id,),
    )
    my_applications = [dict(row) for row in c.fetchall()]
    for application in my_applications:
        team = get_scheduled_team_for_opening(application["opening_date"], application["shift"], application.get("slot_label", ""))
        team_meta = get_team_metadata(team)
        application["team_label"] = f"{team_meta['label']} {team_meta['shift_label']}" if team else application["shift"].title()
    conn.close()

    return render_template(
        "overtime.html",
        openings=openings,
        applied_ids=applied_ids,
        my_applications=my_applications,
        site_scope=site_scope,
        allowed_teams=allowed_teams,
    )


@app.route("/overtime_admin", methods=["GET", "POST"])
@login_required
def overtime_admin():
    if not user_has_rank_access(current_user, SUPERVISOR_RANKS):
        return "Access Denied"

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    site_scope = get_site_scope_for_user(current_user)

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        command_staff_action = {
            "upload_schedule",
            "auto_assign",
            "remove_schedule",
            "approve_application",
            "deny_application",
            "revoke_assignment",
        }
        if action in command_staff_action and not user_has_rank_access(current_user, COMMAND_TEAM_RANKS):
            conn.close()
            flash("Only Training Lieutenants and above can manage overtime assignments.")
            return redirect(url_for("overtime_admin"))

        if action == "upload_schedule":
            period_label = request.form.get("period_label", "").strip()
            site = request.form.get("site", "").strip()
            shift = request.form.get("shift", "").strip().lower()
            schedule_file = request.files.get("schedule_file")
            replace_existing = request.form.get("replace_existing") == "yes"

            if not period_label or not site or site not in SITE_OPTIONS or shift not in {"day", "night"} or not schedule_file:
                conn.close()
                flash("Please provide a period label, valid site, shift, and XLSX schedule file.")
                return redirect(url_for("overtime_admin"))
            if site_scope and site != site_scope:
                conn.close()
                flash("You can only upload schedules for your assigned site.")
                return redirect(url_for("overtime_admin"))

            c.execute(
                """
                SELECT COUNT(*)
                FROM overtime_schedule_uploads
                WHERE period_label = ? AND site = ? AND shift = ? AND original_filename = ?
                """,
                (period_label, site, shift, schedule_file.filename),
            )
            duplicate_name_count = c.fetchone()[0]
            if duplicate_name_count and not replace_existing:
                conn.close()
                flash("A schedule with that same file name is already loaded for this period/site/shift. Remove it first or upload again with Replace checked.")
                return redirect(url_for("overtime_admin"))

            c.execute(
                """
                SELECT COUNT(*)
                FROM overtime_schedule_uploads
                WHERE period_label = ? AND site = ? AND shift = ?
                """,
                (period_label, site, shift),
            )
            existing_period_count = c.fetchone()[0]
            if existing_period_count and not replace_existing:
                conn.close()
                flash("A schedule is already loaded for that period/site/shift. Use Replace Current Schedule or remove the existing one first.")
                return redirect(url_for("overtime_admin"))

            try:
                if existing_period_count:
                    remove_overtime_period_data(c, period_label, site, shift)
                saved_path = save_overtime_schedule(schedule_file, period_label, site, shift)
                parsed = parse_overtime_schedule_workbook(saved_path, site, shift, period_label)
            except Exception as exc:
                conn.close()
                flash(f"Schedule upload failed: {exc}")
                return redirect(url_for("overtime_admin"))

            c.execute(
                """
                INSERT INTO overtime_schedule_uploads (period_label, site, shift, original_filename, file_path, uploaded_by, uploaded_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (period_label, site, shift, schedule_file.filename, saved_path, current_user.id, datetime.now().isoformat()),
            )
            upload_id = c.lastrowid
            replace_overtime_period_data(c, period_label, site, shift, upload_id, parsed)
            log_audit(
                c,
                "overtime_upload",
                upload_id,
                "replaced" if existing_period_count else "uploaded",
                current_user,
                details=f"{'Replaced' if existing_period_count else 'Uploaded'} {shift} OT schedule for {site} / {period_label}",
            )
            conn.commit()
            conn.close()
            flash(f"{'Replaced' if existing_period_count else 'Uploaded'} {shift.title()} schedule for {site} / {period_label}.")
            return redirect(url_for("overtime_admin"))

        if action == "auto_assign":
            period_label = request.form.get("period_label", "").strip()
            site = request.form.get("site", "").strip()
            shift = request.form.get("shift", "").strip().lower()
            if not period_label or not site or not shift:
                conn.close()
                flash("Please choose a valid period/site/shift for auto assignment.")
                return redirect(url_for("overtime_admin"))
            if site_scope and site != site_scope:
                conn.close()
                flash("You can only assign overtime for your assigned site.")
                return redirect(url_for("overtime_admin"))

            result = auto_assign_overtime_period(c, period_label, site, shift, current_user)
            conn.commit()
            conn.close()
            flash(f"Auto-assigned {result['assigned']} openings. {result['unfilled']} remained unfilled.")
            return redirect(url_for("overtime_admin"))

        if action in {"approve_application", "deny_application"}:
            application_id = request.form.get("application_id", "").strip()
            if not application_id:
                conn.close()
                flash("Please choose a valid overtime application.")
                return redirect(url_for("overtime_admin"))

            c.execute(
                """
                SELECT overtime_applications.id AS application_id,
                       overtime_applications.user_id,
                       overtime_applications.status AS application_status,
                       overtime_openings.*,
                       COALESCE(users.first_name, '') AS applicant_first_name,
                       COALESCE(users.last_name, '') AS applicant_last_name,
                       users.username AS applicant_username
                FROM overtime_applications
                JOIN overtime_openings ON overtime_applications.opening_id = overtime_openings.id
                JOIN users ON overtime_applications.user_id = users.id
                WHERE overtime_applications.id = ?
                """,
                (application_id,),
            )
            application = c.fetchone()
            if not application:
                conn.close()
                flash("That overtime application was not found.")
                return redirect(url_for("overtime_admin"))
            if site_scope and application["site"] != site_scope:
                conn.close()
                flash("You can only review overtime for your assigned site.")
                return redirect(url_for("overtime_admin"))

            applicant_name = get_display_name_from_values(
                application["applicant_first_name"],
                application["applicant_last_name"],
                application["applicant_username"],
            )

            if action == "approve_application":
                try:
                    assign_overtime_application(c, application, application["application_id"], current_user, assignment_mode="manual")
                except ValueError as exc:
                    conn.close()
                    flash(str(exc))
                    return redirect(url_for("overtime_admin"))

                conn.commit()
                conn.close()
                flash(f"Approved {applicant_name} for {application['location']} on {application['opening_date']}.")
                return redirect(url_for("overtime_admin"))

            if application["application_status"] != "Applied":
                conn.close()
                flash("Only pending overtime applications can be denied.")
                return redirect(url_for("overtime_admin"))

            c.execute(
                "UPDATE overtime_applications SET status = 'Denied' WHERE id = ?",
                (application["application_id"],),
            )
            create_notification(
                c,
                application["user_id"],
                "Overtime request denied",
                f"Your overtime request for {application['location']} on {application['opening_date']} was denied.",
                category="warning",
                link="/overtime",
            )
            log_audit(
                c,
                "overtime_application",
                application["application_id"],
                "denied",
                current_user,
                target_user_id=application["user_id"],
                details=f"Denied OT request from {applicant_name} for {application['location']} on {application['opening_date']}",
            )
            conn.commit()
            conn.close()
            flash(f"Denied overtime request from {applicant_name}.")
            return redirect(url_for("overtime_admin"))

        if action == "revoke_assignment":
            opening_id = request.form.get("opening_id", "").strip()
            if not opening_id:
                conn.close()
                flash("Please choose a valid assigned overtime opening.")
                return redirect(url_for("overtime_admin"))
            c.execute("SELECT * FROM overtime_openings WHERE id = ?", (opening_id,))
            opening = c.fetchone()
            if not opening:
                conn.close()
                flash("That overtime opening was not found.")
                return redirect(url_for("overtime_admin"))
            if site_scope and opening["site"] != site_scope:
                conn.close()
                flash("You can only revoke overtime for your assigned site.")
                return redirect(url_for("overtime_admin"))
            try:
                revoke_overtime_assignment(c, opening, current_user)
            except ValueError as exc:
                conn.close()
                flash(str(exc))
                return redirect(url_for("overtime_admin"))
            conn.commit()
            conn.close()
            flash(f"Revoked overtime assignment for {opening['location']} on {opening['opening_date']}.")
            return redirect(url_for("overtime_admin"))

        if action == "remove_schedule":
            period_label = request.form.get("period_label", "").strip()
            site = request.form.get("site", "").strip()
            shift = request.form.get("shift", "").strip().lower()
            if not period_label or not site or shift not in {"day", "night"}:
                conn.close()
                flash("Please choose a valid uploaded schedule to remove.")
                return redirect(url_for("overtime_admin"))
            if site_scope and site != site_scope:
                conn.close()
                flash("You can only remove schedules for your assigned site.")
                return redirect(url_for("overtime_admin"))

            remove_overtime_period_data(c, period_label, site, shift)
            log_audit(
                c,
                "overtime_upload",
                None,
                "removed",
                current_user,
                details=f"Removed {shift} OT schedule for {site} / {period_label}",
            )
            conn.commit()
            conn.close()
            flash(f"Removed {shift.title()} schedule for {site} / {period_label}.")
            return redirect(url_for("overtime_admin"))

    periods = get_current_overtime_period(c, site_scope)

    uploads_query = """
        SELECT overtime_schedule_uploads.*, users.username
        FROM overtime_schedule_uploads
        LEFT JOIN users ON overtime_schedule_uploads.uploaded_by = users.id
    """
    upload_params = []
    if site_scope:
        uploads_query += " WHERE overtime_schedule_uploads.site = ?"
        upload_params.append(site_scope)
    uploads_query += " ORDER BY overtime_schedule_uploads.uploaded_at DESC LIMIT 10"
    c.execute(uploads_query, upload_params)
    uploads = c.fetchall()

    openings_query = """
        SELECT overtime_openings.*,
               COALESCE(u.first_name, '') AS assigned_first_name,
               COALESCE(u.last_name, '') AS assigned_last_name,
               u.username AS assigned_username,
               (
                 SELECT COUNT(*)
                 FROM overtime_applications oa
                 WHERE oa.opening_id = overtime_openings.id AND oa.status = 'Applied'
               ) AS applicant_count
        FROM overtime_openings
        LEFT JOIN users u ON overtime_openings.assigned_user_id = u.id
    """
    open_params = []
    if site_scope:
        openings_query += " WHERE overtime_openings.site = ?"
        open_params.append(site_scope)
    openings_query += " ORDER BY overtime_openings.opening_date ASC, overtime_openings.site, overtime_openings.shift, CASE overtime_openings.priority WHEN 'critical' THEN 0 ELSE 1 END, overtime_openings.location"
    c.execute(openings_query, open_params)
    openings = [enrich_overtime_opening_row(row) for row in c.fetchall()]
    openings = sorted(
        openings,
        key=lambda opening: (
            opening["opening_date"],
            0 if str(opening["priority"]).lower() == "critical" else 1,
            opening["site"],
            opening["shift"],
            opening["location"],
            opening["shift_start"],
        ),
    )
    applicant_details = build_overtime_admin_applicant_details(c, openings)
    overtime_summary = {
        "open": sum(1 for row in openings if row["status"] == "Open"),
        "taken": sum(1 for row in openings if row["status"] == "Assigned"),
        "pending_applications": sum(row["applicant_count"] for row in openings),
        "critical_open": sum(1 for row in openings if row["status"] == "Open" and row["priority"] == "critical"),
    }
    conn.close()

    return render_template(
        "overtime_admin.html",
        periods=periods,
        uploads=uploads,
        openings=openings,
        applicant_details=applicant_details,
        overtime_summary=overtime_summary,
        site_options=SITE_OPTIONS,
        site_scope=site_scope,
        can_manage_ot=user_has_rank_access(current_user, COMMAND_TEAM_RANKS),
    )


# ---------------- REQUEST PTO ----------------
@app.route("/request", methods=["GET", "POST"])
@login_required
def request_pto():
    if request.method == "POST":
        start = request.form["start_date"]
        end = request.form["end_date"]
        request_kind = request.form.get("request_kind", "PTO").strip() or "PTO"
        special_type = request.form.get("special_type", "").strip()
        documentation_file = request.files.get("documentation_file")

        from datetime import datetime, timedelta

        try:
            start_date = datetime.strptime(start, "%Y-%m-%d")
            end_date = datetime.strptime(end, "%Y-%m-%d")
        except ValueError:
            flash("Invalid date format.")
            return redirect(url_for("request_pto"))

        if end_date < start_date:
            flash("End date cannot be before start date.")
            return redirect(url_for("request_pto"))

        if request_kind not in {"PTO", "Non-PTO Day"}:
            flash("Please choose a valid request type.")
            return redirect(url_for("request_pto"))

        if special_type and special_type not in SPECIAL_PTO_TYPES:
            flash("Please choose a valid special leave type.")
            return redirect(url_for("request_pto"))

        if special_type:
            if not documentation_file or not (documentation_file.filename or "").strip():
                flash("Documentation image is required for bereavement or jury duty requests.")
                return redirect(url_for("request_pto"))
            if not is_allowed_document(documentation_file.filename):
                flash("Documentation must be a PNG, JPG, JPEG, or WEBP image.")
                return redirect(url_for("request_pto"))

        conn = sqlite3.connect(DB)
        c = conn.cursor()

        documentation_path = ""
        if special_type:
            try:
                documentation_path = save_pto_document(documentation_file, current_user.id)
            except ValueError as exc:
                conn.close()
                flash(str(exc))
                return redirect(url_for("request_pto"))

        current = start_date
        requester_is_command = user_is_command_team(current_user)
        while current <= end_date:
            day_str = current.strftime("%Y-%m-%d")

            if not requester_is_command:
                count = count_approved_team_time_off(c, current_user.team, day_str)
                if count >= MAX_OFFICERS_OFF:
                    conn.close()
                    flash(f"Your team already has {MAX_OFFICERS_OFF} officers approved off on {day_str}. Please submit a shift swap request.")
                    return redirect(url_for("shift_swap_request"))

            current += timedelta(days=1)

        submitted_at = datetime.now().isoformat()
        c.execute(
            """
            INSERT INTO pto_requests (
                user_id, start_date, end_date, status, created_at, request_kind, special_type, documentation_path
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (current_user.id, start, end, "Pending", submitted_at, request_kind, special_type, documentation_path),
        )
        request_id = c.lastrowid

        request_link = f"{get_app_base_url()}/admin"
        create_notification(
            c,
            current_user.id,
            "PTO submitted",
            f"Your {request_kind.lower()} request for {start} through {end} was submitted for review.",
            category="success",
            link="/dashboard",
        )

        command_staff = get_command_staff_users(c, current_user.site)
        create_notifications_for_users(
            c,
            [user[0] for user in command_staff],
            "New PTO request",
            f"{current_user.username} ({current_user.team or 'Unassigned'}) requested {request_kind.lower()} for {start} through {end}. Review is needed."
            + (f" Special leave: {special_type}." if special_type else ""),
            category="action",
            link="/admin",
        )
        log_audit(
            c,
            "pto_request",
            request_id,
            "submitted",
            current_user,
            target_user_id=current_user.id,
            details=f"{request_kind} submitted for {start} through {end} at {current_user.site or 'Unassigned site'}"
            + (f". Special leave: {special_type}" if special_type else "")
            + (". Documentation attached." if documentation_path else ""),
        )

        notification_error = send_notification_email(
            subject=f"Action Needed: PTO Request from {current_user.username}",
            body=(
                f"A new PTO request needs review.\n\n"
                f"Submitted By: {current_user.username}\n"
                f"Rank: {current_user.rank}\n"
                f"Team: {current_user.team or 'Unassigned'}\n"
                f"Site: {current_user.site or 'Unassigned'}\n"
                f"Request Type: {request_kind}\n"
                f"Special Leave: {special_type or 'None'}\n"
                f"Dates Requested: {start} through {end}\n"
                f"Documentation Attached: {'Yes' if documentation_path else 'No'}\n"
                f"Action Needed: Review this request in the PTO admin panel.\n"
                f"Direct Link: {request_link}\n"
                f"Submitted At: {submitted_at}\n"
            ),
            recipients=[user[2] for user in command_staff if user[2]],
        )

        conn.commit()
        conn.close()

        flash("PTO request submitted successfully.")
        if notification_error:
            flash(notification_error)
        return redirect(url_for("dashboard"))

    return render_template("request_pto.html", special_pto_types=SPECIAL_PTO_TYPES)

@app.route("/time_clock", methods=["GET", "POST"])
@login_required
def time_clock():
    flash("Time clock is currently unavailable.")
    return redirect(url_for("dashboard"))

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    now_iso = now.isoformat()
    team = (current_user.team or "").strip()
    today_schedule = get_team_schedule(team, today_str)

    def fetch_today_entry():
        c.execute(
            """
            SELECT *
            FROM time_clock_entries
            WHERE user_id = ? AND clock_date = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (current_user.id, today_str),
        )
        row = c.fetchone()
        return dict(row) if row else None

    if request.method == "POST":
        action = request.form["action"]
        entry = fetch_today_entry()

        if action == "clock_in":
            if not team:
                create_notification(
                    c,
                    current_user.id,
                    "Time clock issue",
                    "Clock-in was blocked because your team is not assigned yet.",
                    category="warning",
                    link="/time_clock",
                )
                conn.commit()
                conn.close()
                flash("An admin must assign your team before you can use the time clock.")
                return redirect(url_for("time_clock"))

            if entry and entry["clock_in"]:
                create_notification(
                    c,
                    current_user.id,
                    "Time clock issue",
                    "Clock-in was skipped because you were already clocked in for today.",
                    category="warning",
                    link="/time_clock",
                )
                conn.commit()
                conn.close()
                flash("You are already clocked in for today.")
                return redirect(url_for("time_clock"))

            c.execute(
                """
                INSERT INTO time_clock_entries (
                    user_id, clock_date, clock_in, clock_out, worked_hours, expected_hours, created_at, updated_at, team, schedule_note,
                    supervisor_note, corrected_by, corrected_at, entry_source, unpaid_break_minutes, paid_lunch_minutes
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    current_user.id,
                    today_str,
                    now_iso,
                    None,
                    None,
                    today_schedule["expected_hours"],
                    now_iso,
                    now_iso,
                    team,
                    today_schedule["schedule_note"],
                    "",
                    "",
                    None,
                    "clock",
                    0,
                    0,
                ),
            )
            entry_id = c.lastrowid
            log_audit(
                c,
                "time_clock",
                entry_id,
                "clock_in",
                current_user,
                target_user_id=current_user.id,
                details=f"Clocked in on {today_str}",
            )
            create_notification(
                c,
                current_user.id,
                "Clocked in",
                f"You clocked in for {today_str}. Full shifts deduct two unpaid 20-minute breaks; half shifts deduct one.",
                category="success",
                link="/time_clock",
            )
            conn.commit()
            conn.close()
            flash(f"Clocked in for {today_str}.")
            return redirect(url_for("time_clock"))

        if not entry or not entry["clock_in"]:
            create_notification(
                c,
                current_user.id,
                "Time clock issue",
                "This action was blocked because no clock-in was found for today.",
                category="warning",
                link="/time_clock",
            )
            conn.commit()
            conn.close()
            flash("You must clock in before using break, lunch, or clock-out actions.")
            return redirect(url_for("time_clock"))

        if entry["clock_out"]:
            create_notification(
                c,
                current_user.id,
                "Time clock issue",
                "This action was blocked because today's shift is already closed.",
                category="warning",
                link="/time_clock",
            )
            conn.commit()
            conn.close()
            flash("Today's shift is already closed.")
            return redirect(url_for("time_clock"))

        active_segment = get_active_time_segment(entry)
        allowed_breaks = get_allowed_unpaid_breaks(entry["expected_hours"])
        flash_message = ""

        if action == "start_break_1":
            if active_segment:
                flash_message = "Finish your current break or lunch before starting another."
            elif entry["break1_start"]:
                flash_message = "Unpaid break 1 has already been started."
            elif allowed_breaks < 1:
                flash_message = "This shift does not have an unpaid break allowance."
            else:
                c.execute("UPDATE time_clock_entries SET break1_start = ?, updated_at = ? WHERE id = ?", (now_iso, now_iso, entry["id"]))
                log_audit(c, "time_clock", entry["id"], "start_break_1", current_user, target_user_id=current_user.id, details=f"Started unpaid break 1 on {today_str}")
                flash_message = "Unpaid break 1 started."

        elif action == "end_break_1":
            if not entry["break1_start"] or entry["break1_end"]:
                flash_message = "Unpaid break 1 has not been started."
            else:
                c.execute("UPDATE time_clock_entries SET break1_end = ?, updated_at = ? WHERE id = ?", (now_iso, now_iso, entry["id"]))
                log_audit(c, "time_clock", entry["id"], "end_break_1", current_user, target_user_id=current_user.id, details=f"Ended unpaid break 1 on {today_str}")
                flash_message = "Unpaid break 1 ended."

        elif action == "start_break_2":
            if active_segment:
                flash_message = "Finish your current break or lunch before starting another."
            elif allowed_breaks < 2:
                flash_message = "This shift only allows one unpaid 20-minute break."
            elif not entry["break1_end"]:
                flash_message = "Finish unpaid break 1 before starting unpaid break 2."
            elif entry["break2_start"]:
                flash_message = "Unpaid break 2 has already been started."
            else:
                c.execute("UPDATE time_clock_entries SET break2_start = ?, updated_at = ? WHERE id = ?", (now_iso, now_iso, entry["id"]))
                log_audit(c, "time_clock", entry["id"], "start_break_2", current_user, target_user_id=current_user.id, details=f"Started unpaid break 2 on {today_str}")
                flash_message = "Unpaid break 2 started."

        elif action == "end_break_2":
            if not entry["break2_start"] or entry["break2_end"]:
                flash_message = "Unpaid break 2 has not been started."
            else:
                c.execute("UPDATE time_clock_entries SET break2_end = ?, updated_at = ? WHERE id = ?", (now_iso, now_iso, entry["id"]))
                log_audit(c, "time_clock", entry["id"], "end_break_2", current_user, target_user_id=current_user.id, details=f"Ended unpaid break 2 on {today_str}")
                flash_message = "Unpaid break 2 ended."

        elif action == "start_lunch":
            if active_segment:
                flash_message = "Finish your current break or lunch before starting another."
            elif entry["lunch_start"]:
                flash_message = "Paid lunch has already been started."
            else:
                c.execute("UPDATE time_clock_entries SET lunch_start = ?, updated_at = ? WHERE id = ?", (now_iso, now_iso, entry["id"]))
                log_audit(c, "time_clock", entry["id"], "start_lunch", current_user, target_user_id=current_user.id, details=f"Started paid lunch on {today_str}")
                flash_message = "Paid lunch started."

        elif action == "end_lunch":
            if not entry["lunch_start"] or entry["lunch_end"]:
                flash_message = "Paid lunch has not been started."
            else:
                c.execute("UPDATE time_clock_entries SET lunch_end = ?, updated_at = ? WHERE id = ?", (now_iso, now_iso, entry["id"]))
                log_audit(c, "time_clock", entry["id"], "end_lunch", current_user, target_user_id=current_user.id, details=f"Ended paid lunch on {today_str}")
                flash_message = "Paid lunch ended."

        elif action == "clock_out":
            if active_segment:
                flash_message = "End your current break or lunch before clocking out."
            else:
                c.execute("UPDATE time_clock_entries SET clock_out = ?, updated_at = ? WHERE id = ?", (now_iso, now_iso, entry["id"]))
                summary = refresh_time_clock_totals(c, entry["id"])
                worked_hours = summary["worked_hours"] if summary else 0
                unpaid_break_minutes = summary["unpaid_break_minutes"] if summary else 0
                paid_lunch_minutes = summary["paid_lunch_minutes"] if summary else 0
                create_notification(
                    c,
                    current_user.id,
                    "Clocked out",
                    f"You clocked out for {today_str}. Worked hours: {worked_hours:.2f}. Unpaid breaks tracked: {unpaid_break_minutes:.0f} minutes. Paid lunch tracked: {paid_lunch_minutes:.0f} minutes.",
                    category="info",
                    link="/time_clock",
                )
                log_audit(
                    c,
                    "time_clock",
                    entry["id"],
                    "clock_out",
                    current_user,
                    target_user_id=current_user.id,
                    details=f"Clocked out on {today_str} with {worked_hours:.2f} hours after {unpaid_break_minutes:.0f} unpaid break minutes",
                )
                conn.commit()
                conn.close()
                flash(f"Clocked out. Worked hours: {worked_hours:.2f}")
                return redirect(url_for("time_clock"))

        else:
            conn.close()
            flash("Unknown time clock action.")
            return redirect(url_for("time_clock"))

        refresh_time_clock_totals(c, entry["id"])
        conn.commit()
        conn.close()
        flash(flash_message)
        return redirect(url_for("time_clock"))

    today_entry = fetch_today_entry()
    today_entry_summary = get_time_entry_summary(today_entry, reference_iso=now_iso) if today_entry else get_time_entry_summary(None)

    c.execute(
        """
        SELECT *
        FROM time_clock_entries
        WHERE user_id = ?
        ORDER BY clock_date DESC, id DESC
        LIMIT 14
        """,
        (current_user.id,),
    )
    recent_entries_raw = c.fetchall()
    recent_entries = []
    for row in recent_entries_raw:
        entry = dict(row)
        entry.update(get_time_entry_summary(entry))
        recent_entries.append(entry)

    conn.close()
    return render_template(
        "time_clock.html",
        today_entry=today_entry,
        today_entry_summary=today_entry_summary,
        recent_entries=recent_entries,
        today=today_str,
        today_schedule=today_schedule,
        team_meta=get_team_metadata(current_user.team),
        unpaid_break_minutes=UNPAID_BREAK_MINUTES,
        paid_lunch_minutes=PAID_LUNCH_MINUTES,
    )



@app.route("/timecard_review", methods=["GET", "POST"])
@login_required
def timecard_review():
    flash("Timecard review is currently unavailable.")
    return redirect(url_for("dashboard"))

    if not user_has_rank_access(current_user, SUPERVISOR_RANKS):
        return "Access Denied"

    site_scope = get_site_scope_for_user(current_user)
    default_start, default_end = get_pay_period_bounds()

    if request.method == "POST":
        conn = sqlite3.connect(DB)
        c = conn.cursor()
        action = request.form.get("action", "").strip()
        filter_start = request.form.get("start_date_filter", default_start)
        filter_end = request.form.get("end_date_filter", default_end)
        filter_user = request.form.get("selected_user_id_filter", "")

        if action == "update_entry":
            entry_id = request.form.get("entry_id", "").strip()
            clock_in = parse_datetime_local(request.form.get("clock_in", ""))
            clock_out = parse_datetime_local(request.form.get("clock_out", ""))
            expected_hours_raw = request.form.get("expected_hours", "").strip()
            supervisor_note = request.form.get("supervisor_note", "").strip()

            c.execute(
                """
                SELECT time_clock_entries.id, time_clock_entries.user_id, time_clock_entries.clock_date, time_clock_entries.team, time_clock_entries.schedule_note,
                       COALESCE(time_clock_entries.unpaid_break_minutes, 0), COALESCE(time_clock_entries.paid_lunch_minutes, 0), users.site
                FROM time_clock_entries
                JOIN users ON time_clock_entries.user_id = users.id
                WHERE time_clock_entries.id = ?
                """,
                (entry_id,),
            )
            entry = c.fetchone()

            if not entry:
                conn.close()
                flash("Time entry not found.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            if site_scope and entry[7] != site_scope:
                conn.close()
                flash("You can only correct time entries for your assigned site.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            if clock_out and not clock_in:
                conn.close()
                flash("Clock-out cannot be set without a clock-in.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            if clock_in and clock_out and datetime.fromisoformat(clock_out) < datetime.fromisoformat(clock_in):
                conn.close()
                flash("Clock-out cannot be earlier than clock-in.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            schedule = get_team_schedule(entry[3], entry[2])
            expected_hours = schedule["expected_hours"] if expected_hours_raw == "" else float(expected_hours_raw)
            worked_hours = calculate_worked_hours(clock_in, clock_out, entry[5]) if clock_in and clock_out else None
            corrected_at = datetime.now().isoformat()

            c.execute(
                """
                UPDATE time_clock_entries
                SET clock_in = ?, clock_out = ?, worked_hours = ?, expected_hours = ?, supervisor_note = ?,
                    corrected_by = ?, corrected_at = ?, entry_source = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    clock_in,
                    clock_out,
                    worked_hours,
                    expected_hours,
                    supervisor_note,
                    current_user.username,
                    corrected_at,
                    "manual",
                    corrected_at,
                    entry_id,
                ),
            )
            create_notification(
                c,
                entry[1],
                "Time entry corrected",
                f"Your time entry for {entry[2]} was reviewed and corrected by {current_user.username}.",
                category="info",
                link="/time_clock",
            )
            log_audit(
                c,
                "time_clock",
                int(entry_id),
                "corrected",
                current_user,
                target_user_id=entry[1],
                details=f"Time entry corrected for {entry[2]}. Note: {supervisor_note or 'No note provided'}",
            )
            conn.commit()
            conn.close()
            flash("Time entry updated.")
            return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

        if action == "create_missing_entry":
            user_id = request.form.get("missing_user_id", "").strip()
            clock_date = request.form.get("missing_clock_date", "").strip()
            clock_in = parse_datetime_local(request.form.get("missing_clock_in", ""))
            clock_out = parse_datetime_local(request.form.get("missing_clock_out", ""))
            expected_hours_raw = request.form.get("missing_expected_hours", "").strip()
            supervisor_note = request.form.get("missing_note", "").strip()

            if not user_id or not clock_date:
                conn.close()
                flash("Employee and date are required to resolve a missing punch.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            if clock_out and not clock_in:
                conn.close()
                flash("Clock-out cannot be set without a clock-in.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            if clock_in and clock_out and datetime.fromisoformat(clock_out) < datetime.fromisoformat(clock_in):
                conn.close()
                flash("Clock-out cannot be earlier than clock-in.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            user_record = get_user_record(c, user_id)
            if not user_record:
                conn.close()
                flash("Employee not found.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            if site_scope and user_record["site"] != site_scope:
                conn.close()
                flash("You can only resolve missing punches for your assigned site.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            c.execute(
                """
                SELECT id
                FROM time_clock_entries
                WHERE user_id = ? AND clock_date = ?
                """,
                (user_id, clock_date),
            )
            existing = c.fetchone()
            if existing:
                conn.close()
                flash("A time entry already exists for that employee and date. Use the correction form instead.")
                return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

            schedule = get_team_schedule(user_record["team"], clock_date)
            expected_hours = schedule["expected_hours"] if expected_hours_raw == "" else float(expected_hours_raw)
            worked_hours = calculate_worked_hours(clock_in, clock_out) if clock_in and clock_out else None
            now_iso = datetime.now().isoformat()

            c.execute(
                """
                INSERT INTO time_clock_entries (
                    user_id, clock_date, clock_in, clock_out, worked_hours, expected_hours, created_at, updated_at,
                    team, schedule_note, supervisor_note, corrected_by, corrected_at, entry_source
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(user_id),
                    clock_date,
                    clock_in,
                    clock_out,
                    worked_hours,
                    expected_hours,
                    now_iso,
                    now_iso,
                    user_record["team"],
                    schedule["schedule_note"],
                    supervisor_note,
                    current_user.username,
                    now_iso,
                    "manual",
                ),
            )
            entry_id = c.lastrowid
            create_notification(
                c,
                int(user_id),
                "Missing punch resolved",
                f"A supervisor created a time entry for {clock_date}. Review your time clock if needed.",
                category="info",
                link="/time_clock",
            )
            log_audit(
                c,
                "time_clock",
                entry_id,
                "missing_punch_resolved",
                current_user,
                target_user_id=int(user_id),
                details=f"Manual time entry created for {clock_date}. Note: {supervisor_note or 'No note provided'}",
            )
            conn.commit()
            conn.close()
            flash("Missing punch resolved.")
            return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

        conn.close()
        flash("Unknown timecard action.")
        return redirect(url_for("timecard_review", start_date=filter_start, end_date=filter_end, user_id=filter_user))

    start_date = request.args.get("start_date", default_start)
    end_date = request.args.get("end_date", default_end)
    selected_user_id = request.args.get("user_id", "")

    context = build_timecard_review_context(start_date, end_date, selected_user_id, site_scope=site_scope)
    return render_template(
        "timecard_review.html",
        start_date=start_date,
        end_date=end_date,
        selected_user_id=selected_user_id,
        pay_period_start=default_start,
        pay_period_end=default_end,
        **context,
    )


@app.route("/timecard_review/export")
@login_required
def timecard_review_export():
    flash("Timecard export is currently unavailable.")
    return redirect(url_for("dashboard"))

    if not user_has_rank_access(current_user, SUPERVISOR_RANKS):
        return "Access Denied"

    site_scope = get_site_scope_for_user(current_user)
    default_start, default_end = get_pay_period_bounds()
    start_date = request.args.get("start_date", default_start)
    end_date = request.args.get("end_date", default_end)
    selected_user_id = request.args.get("user_id", "")

    context = build_timecard_review_context(start_date, end_date, selected_user_id, site_scope=site_scope)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "Employee",
            "Rank",
            "Team",
            "Shift",
            "Expected Hours",
            "Worked Hours",
            "Variance",
            "Unpaid Break Minutes",
            "Paid Lunch Minutes",
            "Missing Punches",
            "Open Shifts",
            "Unscheduled Punches",
            "Late Punches",
            "Exception Count",
        ]
    )
    for summary in context["employee_summaries"]:
        writer.writerow(
            [
                summary["username"],
                summary["rank"],
                summary["team_meta"]["label"],
                summary["team_meta"]["shift_label"],
                summary["expected_total"],
                summary["worked_total"],
                summary["variance"],
                summary["unpaid_break_minutes_total"],
                summary["paid_lunch_minutes_total"],
                summary["missing_punches"],
                summary["open_shifts"],
                summary["unscheduled_punches"],
                summary["late_punches"],
                summary["exception_count"],
            ]
        )

    filename = f"timecard_summary_{start_date}_to_{end_date}.csv"
    return Response(
        buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )




@app.route("/shift_swap", methods=["GET", "POST"])
@login_required
def shift_swap_request():
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    if request.method == "POST":
        swap_with_user_id = request.form["swap_with_user_id"].strip()
        start = request.form.get("taken_shift_date", request.form.get("start_date", "")).strip()
        end = request.form.get("requester_off_date", request.form.get("end_date", "")).strip()
        reason = request.form["reason"].strip()

        try:
            datetime.strptime(start, "%Y-%m-%d")
            datetime.strptime(end, "%Y-%m-%d")
        except ValueError:
            conn.close()
            flash("Please enter valid swap dates.")
            return redirect(url_for("shift_swap_request"))

        if not reason:
            conn.close()
            flash("Please provide a reason for the shift swap request.")
            return redirect(url_for("shift_swap_request"))

        c.execute("SELECT id, username, site FROM users WHERE id=?", (swap_with_user_id,))
        swap_user = c.fetchone()

        if not swap_user:
            conn.close()
            flash("Please choose a valid employee to swap with.")
            return redirect(url_for("shift_swap_request"))

        if int(swap_with_user_id) == current_user.id:
            conn.close()
            flash("You cannot submit a shift swap with yourself.")
            return redirect(url_for("shift_swap_request"))

        if swap_user[2] != current_user.site:
            conn.close()
            flash("Shift swaps must stay within the same site.")
            return redirect(url_for("shift_swap_request"))

        submitted_at = datetime.now().isoformat()
        c.execute(
            """
            INSERT INTO shift_swaps (
                requesting_user_id,
                swap_with_user_id,
                start_date,
                end_date,
                reason,
                status,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                current_user.id,
                int(swap_with_user_id),
                start,
                end,
                reason,
                "Pending",
                submitted_at,
            ),
        )
        swap_id = c.lastrowid

        create_notification(
            c,
            current_user.id,
            "Shift swap submitted",
            f"Your shift swap request for {start} and {end} was submitted for review.",
            category="success",
            link="/dashboard",
        )

        command_staff = get_command_staff_users(c, current_user.site)
        create_notifications_for_users(
            c,
            [user[0] for user in command_staff],
            "New shift swap request",
            f"{current_user.username} needs a shift swap reviewed for {start} and {end}.",
            category="action",
            link="/shift_swap_admin",
        )
        log_audit(
            c,
            "shift_swap",
            swap_id,
            "submitted",
            current_user,
            target_user_id=current_user.id,
            details=f"Shift swap submitted for {start} and {end} at {current_user.site or 'Unassigned site'}",
        )

        notification_error = send_notification_email(
            subject=f"Action Needed: Shift Swap from {current_user.username}",
            body=(
                f"A new shift swap request needs review.\n\n"
                f"Submitted By: {current_user.username}\n"
                f"Rank: {current_user.rank}\n"
                f"Team: {current_user.team or 'Unassigned'}\n"
                f"Site: {current_user.site or 'Unassigned'}\n"
                f"Swap Partner: {swap_user[1]}\n"
                f"Shift Being Picked Up: {start}\n"
                f"Day Requested Off: {end}\n"
                f"Reason: {reason}\n"
                f"Action Needed: Review this request in the shift swap panel.\n"
                f"Direct Link: {get_app_base_url()}/shift_swap_admin\n"
                f"Submitted At: {submitted_at}\n"
            ),
            recipients=[user[2] for user in command_staff if user[2]],
        )

        conn.commit()
        conn.close()
        flash("Shift swap request submitted successfully.")
        if notification_error:
            flash(notification_error)
        return redirect(url_for("dashboard"))

    c.execute("SELECT id, username, rank FROM users WHERE id != ? AND site = ? ORDER BY username", (current_user.id, current_user.site))
    users = c.fetchall()
    conn.close()
    return render_template("shift_swap.html", users=users)


@app.route("/shift_swap_admin", methods=["GET", "POST"])
@login_required
def shift_swap_admin():
    if not user_has_rank_access(current_user, COMMAND_REVIEW_RANKS):
        return "Access Denied"

    site_scope = get_site_scope_for_user(current_user)
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    if request.method == "POST":
        swap_id = request.form["swap_id"]
        action = request.form["action"].strip().lower()
        decision_note = request.form.get("decision_note", "").strip()

        c.execute(
            """
            SELECT shift_swaps.id, shift_swaps.requesting_user_id, shift_swaps.status, shift_swaps.start_date, shift_swaps.end_date, shift_swaps.reason,
                   COALESCE(shift_swaps.review_note, ''), requester.site
            FROM shift_swaps
            JOIN users AS requester ON shift_swaps.requesting_user_id = requester.id
            WHERE shift_swaps.id=?
            """,
            (swap_id,),
        )
        swap = c.fetchone()

        if not swap:
            conn.close()
            flash("Shift swap request not found.")
            return redirect(url_for("shift_swap_admin"))

        if site_scope and swap[7] != site_scope:
            conn.close()
            flash("You can only review shift swaps for your assigned site.")
            return redirect(url_for("shift_swap_admin"))

        if action not in ["approved", "denied"]:
            conn.close()
            flash(f"Unexpected action value: {action!r}")
            return redirect(url_for("shift_swap_admin"))

        new_status = "Approved" if action == "approved" else "Denied"
        c.execute(
            """
            UPDATE shift_swaps
            SET status = ?, approved_by = ?, review_note = ?
            WHERE id = ?
            """,
            (new_status, current_user.username, decision_note, swap_id),
        )
        create_notification(
            c,
            swap[1],
            f"Shift swap {new_status.lower()}",
            f"Your shift swap request was {new_status.lower()} by {current_user.username}."
            + (f" Note: {decision_note}" if decision_note else ""),
            category="info",
            link="/dashboard",
        )
        log_audit(
            c,
            "shift_swap",
            int(swap_id),
            new_status.lower(),
            current_user,
            target_user_id=swap[1],
            details=f"Shift swap was {new_status.lower()} at {swap[7] or 'Unassigned site'}"
            + (f". Note: {decision_note}" if decision_note else ""),
        )
        email_error = send_user_email(
            c,
            swap[1],
            subject=f"Shift Swap {new_status}: {swap[3]} / {swap[4]}",
            body=(
                f"Your shift swap request has been {new_status.lower()}.\n\n"
                f"Reviewed By: {current_user.username}\n"
                f"Shift Being Picked Up: {swap[3]}\n"
                f"Day Requested Off: {swap[4]}\n"
                f"Reason: {swap[5]}\n"
                f"Supervisor Note: {decision_note or 'None provided'}\n"
                f"View Details: {get_app_base_url()}/dashboard\n"
            ),
        )
        conn.commit()
        conn.close()
        flash(f"Shift swap request {new_status.lower()}.")
        if email_error:
            flash(email_error)
        return redirect(url_for("shift_swap_admin"))

    query = """
        SELECT
            shift_swaps.id,
            shift_swaps.requesting_user_id,
            shift_swaps.swap_with_user_id,
            shift_swaps.start_date,
            shift_swaps.end_date,
            shift_swaps.reason,
            shift_swaps.status,
            shift_swaps.approved_by,
            shift_swaps.created_at,
            COALESCE(shift_swaps.review_note, ''),
            requester.username,
            swap_user.username,
            requester.site
        FROM shift_swaps
        JOIN users AS requester ON shift_swaps.requesting_user_id = requester.id
        LEFT JOIN users AS swap_user ON shift_swaps.swap_with_user_id = swap_user.id
    """
    params = []
    if site_scope:
        query += " WHERE requester.site = ?"
        params.append(site_scope)
    query += """
        ORDER BY
            CASE shift_swaps.status
                WHEN 'Pending' THEN 0
                WHEN 'Approved' THEN 1
                WHEN 'Denied' THEN 2
                ELSE 3
            END,
            shift_swaps.created_at DESC
    """
    c.execute(query, params)
    shift_swaps = c.fetchall()

    conn.close()
    return render_template("shift_swap_admin.html", shift_swaps=shift_swaps)

# ---------------- ADMIN ----------------

@app.route("/pto_document/<int:req_id>")
@login_required
def pto_document(req_id):
    if not user_has_rank_access(current_user, COMMAND_REVIEW_RANKS):
        return "Access Denied"

    site_scope = get_site_scope_for_user(current_user)
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute(
        """
        SELECT pto_requests.documentation_path, users.site
        FROM pto_requests
        JOIN users ON pto_requests.user_id = users.id
        WHERE pto_requests.id = ?
        """,
        (req_id,),
    )
    row = c.fetchone()
    conn.close()

    if not row or not row[0]:
        flash("Documentation file was not found for that request.")
        return redirect(url_for("admin"))

    if site_scope and row[1] != site_scope:
        return "Access Denied"

    file_path = Path(row[0])
    if not file_path.exists():
        flash("Documentation file is missing from storage.")
        return redirect(url_for("admin"))

    return send_file(file_path)


@app.route("/admin", methods=["GET", "POST"])
@login_required
def admin():
    if not user_has_rank_access(current_user, COMMAND_REVIEW_RANKS):
        return "Access Denied"

    site_scope = get_site_scope_for_user(current_user)
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    if request.method == "POST":
        req_id = request.form.get("req_id", "")
        action = request.form["action"]
        decision_note = request.form.get("decision_note", "").strip()
        force_override = request.form.get("force_override") == "yes" and current_user.role == "admin"

        if action.strip().lower() in {"approved", "denied", "clear_single"}:
            c.execute(
                """
                SELECT pto_requests.id, pto_requests.user_id, pto_requests.start_date, pto_requests.end_date, pto_requests.status,
                       pto_requests.request_kind, pto_requests.special_type, pto_requests.documentation_path,
                       users.site, users.rank, users.team, users.role
                FROM pto_requests
                JOIN users ON pto_requests.user_id = users.id
                WHERE pto_requests.id = ?
                """,
                (req_id,),
            )
            request_data = c.fetchone()

            if not request_data:
                conn.close()
                flash("Request not found.")
                return redirect(url_for("admin"))

            if site_scope and request_data[8] != site_scope:
                conn.close()
                flash("You can only manage PTO requests for your assigned site.")
                return redirect(url_for("admin"))

        if action.strip().lower() == "approved":
            _, request_user_id, start, end, _, request_kind, special_type, documentation_path, request_site, request_rank, request_team, request_role = request_data
            start_date = datetime.strptime(start, "%Y-%m-%d")
            end_date = datetime.strptime(end, "%Y-%m-%d")

            current = start_date
            requester_is_command = rank_team_is_command(request_rank, request_team, request_role)
            while current <= end_date:
                day_str = current.strftime("%Y-%m-%d")
                if not requester_is_command and not force_override:
                    count = count_approved_team_time_off(c, request_team, day_str, exclude_request_id=req_id)
                    if count >= MAX_OFFICERS_OFF:
                        conn.close()
                        flash(f"Cannot approve request. Team {request_team or 'Unassigned'} already has {MAX_OFFICERS_OFF} officers off on {day_str}.")
                        return redirect(url_for("admin"))

                current += timedelta(days=1)

            c.execute(
                "UPDATE pto_requests SET status = 'Approved', admin_note = ?, override_approved = ? WHERE id = ?",
                (decision_note, 1 if force_override else 0, req_id),
            )
            create_notification(
                c,
                request_user_id,
                "PTO approved",
                f"Your PTO request for {start} through {end} was approved by {current_user.username}."
                + (f" Note: {decision_note}" if decision_note else ""),
                category="success",
                link="/dashboard",
            )
            log_audit(
                c,
                "pto_request",
                int(req_id),
                "approved",
                current_user,
                target_user_id=request_user_id,
                details=f"{request_kind} approved for {start} through {end} at {request_site or 'Unassigned site'}"
                + (f". Special leave: {special_type}" if special_type else "")
                + (". Documentation attached." if documentation_path else "")
                + (". Admin override used." if force_override else "")
                + (f". Note: {decision_note}" if decision_note else ""),
            )
            email_error = send_user_email(
                c,
                request_user_id,
                subject=f"{request_kind} Approved: {start} through {end}",
                body=(
                    f"Your {request_kind.lower()} request has been approved.\n\n"
                    f"Reviewed By: {current_user.username}\n"
                    f"Dates Approved: {start} through {end}\n"
                    f"Special Leave: {special_type or 'None'}\n"
                    f"Admin Override: {'Yes' if force_override else 'No'}\n"
                    f"Supervisor Note: {decision_note or 'None provided'}\n"
                    f"View Details: {get_app_base_url()}/dashboard\n"
                ),
            )
            conn.commit()
            conn.close()
            flash("Request approved.")
            if email_error:
                flash(email_error)
            return redirect(url_for("admin"))

        elif action.strip().lower() == "denied":
            _, request_user_id, start, end, _, request_kind, special_type, documentation_path, request_site, request_rank, request_team, request_role = request_data
            c.execute(
                "UPDATE pto_requests SET status = 'Denied', admin_note = ?, override_approved = 0 WHERE id = ?",
                (decision_note, req_id),
            )
            create_notification(
                c,
                request_user_id,
                "PTO denied",
                f"Your PTO request for {start} through {end} was denied by {current_user.username}."
                + (f" Note: {decision_note}" if decision_note else ""),
                category="warning",
                link="/dashboard",
            )
            log_audit(
                c,
                "pto_request",
                int(req_id),
                "denied",
                current_user,
                target_user_id=request_user_id,
                details=f"{request_kind} denied for {start} through {end} at {request_site or 'Unassigned site'}"
                + (f". Special leave: {special_type}" if special_type else "")
                + (". Documentation attached." if documentation_path else "")
                + (f". Note: {decision_note}" if decision_note else ""),
            )
            email_error = send_user_email(
                c,
                request_user_id,
                subject=f"{request_kind} Denied: {start} through {end}",
                body=(
                    f"Your {request_kind.lower()} request has been denied.\n\n"
                    f"Reviewed By: {current_user.username}\n"
                    f"Dates Requested: {start} through {end}\n"
                    f"Special Leave: {special_type or 'None'}\n"
                    f"Supervisor Note: {decision_note or 'None provided'}\n"
                    f"View Details: {get_app_base_url()}/dashboard\n"
                ),
            )
            conn.commit()
            conn.close()
            flash("Request denied.")
            if email_error:
                flash(email_error)
            return redirect(url_for("admin"))

        elif action == "clear_single":
            _, request_user_id, start, end, current_status, request_kind, special_type, documentation_path, request_site, request_rank, request_team, request_role = request_data
            c.execute("DELETE FROM pto_requests WHERE id = ?", (req_id,))
            if documentation_path:
                try:
                    Path(documentation_path).unlink(missing_ok=True)
                except OSError:
                    pass
            log_audit(
                c,
                "pto_request",
                int(req_id),
                "cleared_single",
                current_user,
                target_user_id=request_user_id,
                details=f"Cleared single {request_kind} request ({current_status}) for {start} through {end} at {request_site or 'Unassigned site'}"
                + (f". Special leave: {special_type}" if special_type else ""),
            )
            conn.commit()
            conn.close()
            flash("PTO request cleared.")
            return redirect(url_for("admin"))

        elif action == "clear_decided":
            clear_shift_swaps = request.form.get("clear_shift_swaps") == "yes"
            params = ["Approved", "Denied"]
            pto_sql = """
                DELETE FROM pto_requests
                WHERE id IN (
                    SELECT pto_requests.id
                    FROM pto_requests
                    JOIN users ON pto_requests.user_id = users.id
                    WHERE pto_requests.status IN (?, ?)
            """
            if site_scope:
                pto_sql += " AND users.site = ?"
                params.append(site_scope)
            pto_sql += ")"
            c.execute(pto_sql, params)
            cleared_pto = c.rowcount

            cleared_swaps = 0
            if clear_shift_swaps:
                swap_params = ["Approved", "Denied"]
                swap_sql = """
                    DELETE FROM shift_swaps
                    WHERE id IN (
                        SELECT shift_swaps.id
                        FROM shift_swaps
                        JOIN users AS requester ON shift_swaps.requesting_user_id = requester.id
                        WHERE shift_swaps.status IN (?, ?)
                """
                if site_scope:
                    swap_sql += " AND requester.site = ?"
                    swap_params.append(site_scope)
                swap_sql += ")"
                c.execute(swap_sql, swap_params)
                cleared_swaps = c.rowcount

            log_audit(
                c,
                "pto_request",
                None,
                "cleared_decided",
                current_user,
                details=f"Cleared {cleared_pto} approved/denied PTO requests and {cleared_swaps} decided shift swaps for {site_scope or 'all sites'}",
            )
            conn.commit()
            conn.close()
            flash(f"Cleared {cleared_pto} approved/denied PTO requests.")
            if clear_shift_swaps:
                flash(f"Cleared {cleared_swaps} approved/denied shift swaps.")
            return redirect(url_for("admin"))

        else:
            conn.close()
            flash(f"Unexpected action value: {action!r}")
            return redirect(url_for("admin"))

    query = """
        SELECT pto_requests.id, pto_requests.user_id, pto_requests.start_date, pto_requests.end_date, pto_requests.status,
               pto_requests.admin_note, pto_requests.created_at, pto_requests.request_kind, pto_requests.special_type, pto_requests.documentation_path,
               users.username, users.rank, users.team, users.site, COALESCE(pto_requests.override_approved, 0)
        FROM pto_requests
        JOIN users ON pto_requests.user_id = users.id
    """
    params = []
    if site_scope:
        query += " WHERE users.site = ?"
        params.append(site_scope)
    query += """
        ORDER BY
            CASE pto_requests.status
                WHEN 'Pending' THEN 0
                WHEN 'Approved' THEN 1
                WHEN 'Denied' THEN 2
                ELSE 3
            END,
            pto_requests.created_at DESC
    """
    c.execute(query, params)
    data = c.fetchall()

    conn.close()
    pending_count = len([request for request in data if request[4] == "Pending"])
    approved_count = len([request for request in data if request[4] == "Approved"])
    denied_count = len([request for request in data if request[4] == "Denied"])
    return render_template(
        "admin.html",
        requests=data,
        pending_count=pending_count,
        approved_count=approved_count,
        denied_count=denied_count,
        current_site=current_user.site,
        site_scope=site_scope,
    )

# ---------------- CALENDAR ----------------
@app.route("/calendar")
@login_required
def calendar():
    import calendar as cal
    from datetime import datetime, timedelta

    today = datetime.today()
    site_scope = get_site_scope_for_user(current_user)
    show_calendar_details = current_user.role == "admin" or current_user.rank in ["Lieutenant", "Training Lieutenant", "Captain", "Director"]

    raw_month = request.args.get("month")
    raw_year = request.args.get("year")

    try:
        month = int(raw_month) if raw_month else today.month
    except:
        month = today.month

    try:
        year = int(raw_year) if raw_year else today.year
    except:
        year = today.year

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    query = """
        SELECT pto_requests.start_date, pto_requests.end_date, pto_requests.request_kind,
               users.username, users.rank, users.team, COALESCE(users.first_name, '') AS first_name, COALESCE(users.last_name, '') AS last_name
        FROM pto_requests
        JOIN users ON pto_requests.user_id = users.id
        WHERE pto_requests.status='Approved'
    """
    params = []
    if site_scope:
        query += " AND users.site = ?"
        params.append(site_scope)
    c.execute(query, params)
    pto_data = c.fetchall()
    conn.close()

    calendar_days = {}
    calendar_team_counts = {}
    calendar_people = {}
    rank_order = {"Director": 1, "Captain": 2, "Training Lieutenant": 3, "Lieutenant": 4, "Sergeant": 5, "Officer": 6}

    for row in pto_data:
        start_date = datetime.strptime(row["start_date"], "%Y-%m-%d")
        end_date = datetime.strptime(row["end_date"], "%Y-%m-%d")
        display_name = get_display_name_from_values(row["first_name"], row["last_name"], row["username"])

        current = start_date
        while current <= end_date:
            day_str = current.strftime("%Y-%m-%d")
            team_key = row["team"] or "Unassigned"
            calendar_team_counts.setdefault(day_str, {})
            calendar_team_counts[day_str][team_key] = calendar_team_counts[day_str].get(team_key, 0) + 1
            calendar_days[day_str] = max(calendar_days.get(day_str, 0), calendar_team_counts[day_str][team_key])
            calendar_people.setdefault(day_str, []).append(
                {
                    "display_name": display_name,
                    "rank": row["rank"] or "Officer",
                    "team": team_key,
                    "request_kind": row["request_kind"] or "PTO",
                }
            )
            current += timedelta(days=1)

    for day_str in calendar_people:
        calendar_people[day_str].sort(key=lambda item: (rank_order.get(item["rank"], 99), item["display_name"].lower()))

    month_days = cal.monthcalendar(year, month)
    month_name = cal.month_name[month]
    month_names = {index: cal.month_name[index] for index in range(1, 13)}
    year_options = list(range(today.year - 2, today.year + 4))
    visible_days = sum(1 for day in calendar_days if day.startswith(f"{year}-{month:02d}-"))
    total_approved = sum(calendar_days.get(f"{year}-{month:02d}-{day:02d}", 0) for week in month_days for day in week if day)

    prev_month = month - 1 if month > 1 else 12
    next_month = month + 1 if month < 12 else 1
    prev_year = year if month > 1 else year - 1
    next_year = year if month < 12 else year + 1

    return render_template(
        "calendar.html",
        calendar_days=calendar_days,
        calendar_team_counts=calendar_team_counts,
        calendar_people=calendar_people,
        show_calendar_details=show_calendar_details,
        visible_days=visible_days,
        total_approved=total_approved,
        site_scope=site_scope,
        month_days=month_days,
        month=month,
        month_name=month_name,
        month_names=month_names,
        year=year,
        year_options=year_options,
        prev_month=prev_month,
        prev_year=prev_year,
        next_month=next_month,
        next_year=next_year
    )

@app.route("/manage_users", methods=["GET", "POST"])
@login_required
def manage_users():
    assignment_ranks = {"Training Lieutenant", "Captain", "Director"}
    building_ranks = {"Lieutenant", "Training Lieutenant", "Captain", "Director"}
    can_assign_site_team = current_user.role == "admin" or current_user.rank in assignment_ranks
    can_assign_building = current_user.role == "admin" or current_user.rank in building_ranks
    is_full_admin = current_user.role == "admin"

    if not (can_assign_site_team or can_assign_building):
        return "Access Denied"

    conn = sqlite3.connect(DB)
    c = conn.cursor()

    if request.method == "POST":
        action = request.form["action"]
        user_id = request.form["user_id"]

        admin_only_actions = {"update_role", "update_rank", "clear_pto", "reset_password", "delete_user"}
        site_team_actions = {"update_team", "update_site"}
        building_actions = {"update_building"}
        if action in admin_only_actions and not is_full_admin:
            conn.close()
            flash("Only admins can perform that action.")
            return redirect(url_for("manage_users"))
        if action in site_team_actions and not can_assign_site_team:
            conn.close()
            flash("You do not have permission to update teams or sites.")
            return redirect(url_for("manage_users"))
        if action in building_actions and not can_assign_building:
            conn.close()
            flash("You do not have permission to update buildings.")
            return redirect(url_for("manage_users"))

        if action == "update_rank":
            new_rank = request.form["rank"]
            command_staff_ranks = {"Training Lieutenant", "Captain", "Director"}
            c.execute("SELECT rank, team FROM users WHERE id=?", (user_id,))
            old_rank = c.fetchone()
            old_rank_value = old_rank[0] if old_rank else "Unknown"
            old_team_value = old_rank[1] if old_rank and len(old_rank) > 1 else ""
            updated_team = old_team_value

            if new_rank in command_staff_ranks:
                updated_team = "Command Staff"
            elif old_team_value == "Command Staff":
                updated_team = ""

            c.execute("UPDATE users SET rank=?, team=? WHERE id=?", (new_rank, updated_team, user_id))
            create_notification(
                c,
                user_id,
                "Rank updated",
                f"Your rank was updated to {new_rank} by {current_user.username}.",
                category="info",
                link="/dashboard",
            )
            log_audit(
                c,
                "user",
                int(user_id),
                "rank_changed",
                current_user,
                target_user_id=int(user_id),
                details=f"Rank changed from {old_rank_value} to {new_rank}. Team is now {updated_team or 'Unassigned'}.",
            )
            conn.commit()
            flash("Rank updated successfully.")

        elif action == "update_team":
            new_team = request.form["team"]
            c.execute("SELECT team FROM users WHERE id=?", (user_id,))
            old_team = c.fetchone()
            c.execute("UPDATE users SET team=? WHERE id=?", (new_team, user_id))
            log_audit(
                c,
                "user",
                int(user_id),
                "team_changed",
                current_user,
                target_user_id=int(user_id),
                details=f"Team changed from {(old_team[0] if old_team and old_team[0] else 'Unassigned')} to {new_team or 'Unassigned'}",
            )
            conn.commit()
            flash("Team updated successfully.")

        elif action == "update_site":
            new_site = request.form["site"]
            if new_site not in SITE_OPTIONS:
                flash("Please choose a valid site.")
            else:
                c.execute("SELECT site FROM users WHERE id=?", (user_id,))
                old_site = c.fetchone()
                c.execute("UPDATE users SET site=? WHERE id=?", (new_site, user_id))
                log_audit(
                    c,
                    "user",
                    int(user_id),
                    "site_changed",
                    current_user,
                    target_user_id=int(user_id),
                    details=f"Site changed from {(old_site[0] if old_site and old_site[0] else 'Unassigned')} to {new_site}",
                )
                conn.commit()
                flash("Site updated successfully.")

        elif action == "update_building":
            new_building = request.form["building"]
            if new_building and new_building not in BUILDING_OPTIONS:
                flash("Please choose a valid building.")
            else:
                c.execute("SELECT building FROM users WHERE id=?", (user_id,))
                old_building = c.fetchone()
                c.execute("UPDATE users SET building=? WHERE id=?", (new_building, user_id))
                log_audit(
                    c,
                    "user",
                    int(user_id),
                    "building_changed",
                    current_user,
                    target_user_id=int(user_id),
                    details=f"Building changed from {(old_building[0] if old_building and old_building[0] else 'Unassigned')} to {new_building or 'Unassigned'}",
                )
                conn.commit()
                flash("Building updated successfully.")

        elif action == "update_role":
            new_role = request.form["role"]

            if str(current_user.id) == str(user_id) and new_role != "admin":
                flash("You cannot remove your own admin access.")
            else:
                c.execute("SELECT COUNT(*) FROM users WHERE role='admin'")
                admin_count = c.fetchone()[0]

                c.execute("SELECT role FROM users WHERE id=?", (user_id,))
                result = c.fetchone()

                if result is None:
                    flash("User not found.")
                else:
                    current_role = result[0]

                    if current_role == "admin" and new_role != "admin" and admin_count <= 1:
                        flash("You cannot demote the last admin account.")
                    else:
                        c.execute("UPDATE users SET role=? WHERE id=?", (new_role, user_id))
                        log_audit(
                            c,
                            "user",
                            int(user_id),
                            "role_changed",
                            current_user,
                            target_user_id=int(user_id),
                            details=f"Role changed from {current_role} to {new_role}",
                        )
                        conn.commit()
                        flash("User role updated successfully.")

        elif action == "clear_pto":
            c.execute("DELETE FROM pto_requests WHERE user_id=?", (user_id,))
            log_audit(
                c,
                "pto_request",
                None,
                "cleared",
                current_user,
                target_user_id=int(user_id),
                details="All PTO requests cleared for user",
            )
            conn.commit()
            flash("All PTO requests for that user were removed.")

        elif action == "reset_password":
            new_password = request.form["new_password"].strip()

            if not new_password:
                flash("Password cannot be blank.")
            else:
                hashed_password = generate_password_hash(new_password)
                c.execute("UPDATE users SET password=? WHERE id=?", (hashed_password, user_id))
                log_audit(
                    c,
                    "user",
                    int(user_id),
                    "password_reset",
                    current_user,
                    target_user_id=int(user_id),
                    details="Password reset by admin",
                )
                conn.commit()
                flash("Password reset successfully.")

        elif action == "delete_user":
            if str(current_user.id) == str(user_id):
                flash("You cannot delete your own account.")
            else:
                c.execute("SELECT COUNT(*) FROM users WHERE role='admin'")
                admin_count = c.fetchone()[0]

                c.execute("SELECT role FROM users WHERE id=?", (user_id,))
                result = c.fetchone()

                if result is None:
                    flash("User not found.")
                else:
                    user_role = result[0]

                    if user_role == "admin" and admin_count <= 1:
                        flash("You cannot delete the last admin account.")
                    else:
                        c.execute("DELETE FROM pto_requests WHERE user_id=?", (user_id,))
                        c.execute("DELETE FROM shift_swaps WHERE requesting_user_id=? OR swap_with_user_id=?", (user_id, user_id))
                        c.execute("DELETE FROM time_clock_entries WHERE user_id=?", (user_id,))
                        c.execute("DELETE FROM notifications WHERE user_id=?", (user_id,))
                        c.execute("DELETE FROM users WHERE id=?", (user_id,))
                        log_audit(
                            c,
                            "user",
                            int(user_id),
                            "deleted",
                            current_user,
                            target_user_id=int(user_id),
                            details="User removed from database",
                        )
                        conn.commit()
                        flash("User removed from database.")

        return redirect(url_for("manage_users"))

    c.execute(
        """
        SELECT id, username, role, rank, email, phone, team, site, COALESCE(building, ''), COALESCE(first_name, ''), COALESCE(last_name, '')
        FROM users
        ORDER BY
            site,
            CASE rank
                WHEN 'Director' THEN 1
                WHEN 'Captain' THEN 2
                WHEN 'Training Lieutenant' THEN 3
                WHEN 'Lieutenant' THEN 4
                WHEN 'Sergeant' THEN 5
                WHEN 'Officer' THEN 6
                ELSE 7
            END,
            last_name,
            first_name,
            username
        """
    )
    users = c.fetchall()
    conn.close()

    return render_template(
        "manage_users.html",
        users=users,
        site_options=SITE_OPTIONS,
        building_options=BUILDING_OPTIONS,
        can_assign_site_team=can_assign_site_team,
        can_assign_building=can_assign_building,
        is_full_admin=is_full_admin,
    )
@app.route("/audit_history")
@login_required
def audit_history():
    if not user_has_rank_access(current_user, SUPERVISOR_RANKS):
        return "Access Denied"

    site_scope = get_site_scope_for_user(current_user)
    entity_type = request.args.get("entity_type", "").strip()
    action = request.args.get("action", "").strip()
    target_user_id = request.args.get("target_user_id", "").strip()

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    query = """
        SELECT
            audit_log.*,
            target.username AS target_username,
            COALESCE(target.site, actor.site, '') AS site_name
        FROM audit_log
        LEFT JOIN users AS target ON audit_log.target_user_id = target.id
        LEFT JOIN users AS actor ON audit_log.actor_user_id = actor.id
        WHERE 1 = 1
    """
    params = []

    if site_scope:
        query += " AND COALESCE(target.site, actor.site, '') = ?"
        params.append(site_scope)
    if entity_type:
        query += " AND audit_log.entity_type = ?"
        params.append(entity_type)
    if action:
        query += " AND audit_log.action = ?"
        params.append(action)
    if target_user_id:
        query += " AND audit_log.target_user_id = ?"
        params.append(target_user_id)

    query += " ORDER BY audit_log.created_at DESC LIMIT 250"
    c.execute(query, params)
    entries = [dict(row) for row in c.fetchall()]

    user_query = "SELECT id, username, rank, site FROM users"
    user_params = []
    if site_scope:
        user_query += " WHERE site = ?"
        user_params.append(site_scope)
    user_query += " ORDER BY username"
    c.execute(user_query, user_params)
    users = c.fetchall()
    conn.close()

    entity_types = sorted({entry["entity_type"] for entry in entries if entry["entity_type"]})
    actions = sorted({entry["action"] for entry in entries if entry["action"]})

    return render_template(
        "audit_history.html",
        entries=entries,
        users=users,
        entity_types=entity_types,
        actions=actions,
        selected_entity_type=entity_type,
        selected_action=action,
        selected_target_user_id=target_user_id,
        site_scope=site_scope,
    )

@app.route("/call_list")
@login_required
def call_list():
    allowed_ranks = ["Lieutenant", "Training Lieutenant", "Captain", "Director"]

    if current_user.role != "admin" and current_user.rank not in allowed_ranks:
        return "Access Denied"

    site_scope = get_site_scope_for_user(current_user)
    selected_rank = request.args.get("rank", "").strip()
    rank_options = list(RANK_OPTIONS)

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    query = """
        SELECT username, rank, team, email, phone, site,
               COALESCE(first_name, '') AS first_name, COALESCE(last_name, '') AS last_name
        FROM users
        WHERE 1 = 1
    """
    params = []
    if site_scope:
        query += " AND site = ?"
        params.append(site_scope)
    if selected_rank:
        query += " AND rank = ?"
        params.append(selected_rank)
    c.execute(query, params)
    rows = [dict(row) for row in c.fetchall()]
    conn.close()

    rank_order = {"Director": 6, "Captain": 5, "Training Lieutenant": 4, "Lieutenant": 3, "Sergeant": 2, "Officer": 1}
    for row in rows:
        row["display_name"] = get_display_name_from_values(row["first_name"], row["last_name"], row["username"])

    users = sorted(
        rows,
        key=lambda item: (-rank_order.get(item["rank"], 0), (item["site"] or ""), item["last_name"].lower(), item["first_name"].lower(), item["username"].lower()),
    )
    team_meta = {team: get_team_metadata(team) for team in TEAM_OPTIONS}

    return render_template(
        "call_list.html",
        users=users,
        team_meta=team_meta,
        site_scope=site_scope,
        rank_options=rank_options,
        selected_rank=selected_rank,
    )


@app.route("/command_notice", methods=["GET", "POST"])
@login_required
def command_notice():
    if not user_has_rank_access(current_user, COMMAND_REVIEW_RANKS):
        return "Access Denied"

    conn = sqlite3.connect(DB)
    c = conn.cursor()
    site_scope = get_site_scope_for_user(current_user)
    selected_user_ids = parse_selected_user_ids(request.form.getlist("selected_user_ids")) if request.method == "POST" else []
    recipients_preview = get_broadcast_recipients(c, site_scope=site_scope)
    recipients_by_id = {recipient["id"]: recipient for recipient in recipients_preview}

    if request.method == "POST":
        subject = request.form.get("subject", "").strip()
        message = request.form.get("message", "").strip()
        send_in_app = request.form.get("send_in_app") == "yes"
        if not subject or not message:
            conn.close()
            flash("Please provide both a subject and message.")
            return redirect(url_for("command_notice"))

        if not selected_user_ids:
            conn.close()
            flash("Please select at least one recipient.")
            return redirect(url_for("command_notice"))

        recipients = [recipients_by_id[user_id] for user_id in selected_user_ids if user_id in recipients_by_id]
        emails = [recipient["email"] for recipient in recipients if recipient["email"]]
        if not emails:
            conn.close()
            flash("No recipients with email addresses matched your filters.")
            return redirect(url_for("command_notice"))

        sender_name = get_display_name_from_values(current_user.first_name, current_user.last_name, current_user.username)
        sender_scope = site_scope or "All Sites"
        email_body = (
            f"{message}\n\n"
            f"Sent By: {sender_name}\n"
            f"Rank: {current_user.rank}\n"
            f"Site Scope: {sender_scope}\n"
            f"Sent At: {datetime.now().isoformat()}"
        )
        email_error = send_notification_email(subject, email_body, emails)
        if email_error:
            conn.close()
            flash(email_error)
            return redirect(url_for("command_notice"))

        if send_in_app:
            create_notifications_for_users(
                c,
                [recipient["id"] for recipient in recipients],
                subject,
                message,
                category="info",
                link="/dashboard",
            )

        log_audit(
            c,
            "command_notice",
            None,
            "sent",
            current_user,
            details=f"Sent command notice to {len(recipients)} users for {sender_scope}",
        )
        conn.commit()
        conn.close()
        flash(f"Notice sent to {len(recipients)} users.")
        return redirect(url_for("command_notice"))

    conn.close()

    return render_template(
        "command_notice.html",
        site_scope=site_scope,
        selected_user_ids=selected_user_ids,
        recipients_preview=recipients_preview,
        recipient_count=len(recipients_preview),
    )

# ---------------- ACCOUNT SETTINGS ----------------
@app.route("/account_settings", methods=["GET", "POST"])
@login_required
def account_settings():
    if request.method == "POST":
        action = request.form.get("action", "").strip()
        conn = sqlite3.connect(DB)
        c = conn.cursor()

        if action == "update_profile":
            first_name = request.form.get("first_name", "").strip()
            last_name = request.form.get("last_name", "").strip()
            email = request.form.get("email", "").strip()
            phone = normalize_phone_number(request.form.get("phone", ""))

            if not first_name or not last_name:
                conn.close()
                flash("First and last name are required.")
                return redirect(url_for("account_settings"))

            c.execute(
                "UPDATE users SET first_name = ?, last_name = ?, email = ?, phone = ? WHERE id = ?",
                (first_name, last_name, email, phone, current_user.id),
            )
            conn.commit()
            conn.close()
            flash("Account settings updated.")
            return redirect(url_for("account_settings"))

        if action == "update_password":
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirm_new_password = request.form.get("confirm_new_password", "")

            if not check_password_hash(current_user.password, current_password):
                conn.close()
                flash("Current password is incorrect.")
                return redirect(url_for("account_settings"))

            if not new_password:
                conn.close()
                flash("New password cannot be blank.")
                return redirect(url_for("account_settings"))

            if new_password != confirm_new_password:
                conn.close()
                flash("New passwords do not match.")
                return redirect(url_for("account_settings"))

            c.execute("UPDATE users SET password = ? WHERE id = ?", (generate_password_hash(new_password), current_user.id))
            conn.commit()
            conn.close()
            flash("Password updated successfully.")
            return redirect(url_for("account_settings"))

        conn.close()
        flash("Unknown settings action.")
        return redirect(url_for("account_settings"))

    return render_template("account_settings.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))

# ---------------- CREATE ADMIN ----------------
def create_admin():
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    try:
        c.execute("INSERT INTO users (username, password, role, rank, email, phone, team, site) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (
    "admin",
    generate_password_hash("admin123"),
    "admin",
    "Director",
    "",
    "",
    "",
    ""
))
        conn.commit()
    except:
        pass

    conn.close()

create_admin()

# ---------------- RUN APP ----------------
if __name__ == "__main__":
    app.run(debug=True)
